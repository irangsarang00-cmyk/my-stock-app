import streamlit as st
import io
import os
import re
import zipfile
from datetime import datetime
from PIL import Image, ImageDraw, ImageFont
from pypdf import PdfWriter, PdfReader
import pdfplumber
import pandas as pd
import openpyxl
import logging

logging.getLogger("pdfminer").setLevel(logging.ERROR)

st.set_page_config(page_title="가평 업무 도우미", layout="wide")

tab_3창고, tab_1창고 = st.tabs(["🏭 3창고", "🏢 1창고"])


# =====================================================================
# 3창고 탭
# =====================================================================
with tab_3창고:
    col_left, col_right = st.columns(2)

    with col_left:
        st.markdown("### 📄 [1] 3창고 서류 취합")
        st.markdown("<div style='background-color: #f0f2f5; padding: 15px; border-radius: 10px;'>", unsafe_allow_html=True)

        uploaded_zips = st.file_uploader("📁 ZIP 파일 업로드", type="zip", accept_multiple_files=True, key="zip_uploader")

        if st.button("🚀 실행", type="primary", use_container_width=True, key="run_btn"):
            if not uploaded_zips:
                st.warning("ZIP 파일 선택 필요")
            else:
                with st.spinner("처리중"):
                    try:
                        def zip_sort_key(file_obj):
                            filename = file_obj.name
                            if '글로브' in filename: return 0
                            if '빌리브' in filename: return 1
                            return 2

                        sorted_zips = sorted(uploaded_zips, key=zip_sort_key)
                        date_match = re.search(r'(\d{8}|\d{6}|\d{4})', sorted_zips[0].name)
                        date_str = date_match.group(1) if date_match else datetime.now().strftime("%Y%m%d")

                        all_pdfs_to_merge = []
                        extracted_table_data = []

                        for z_file in sorted_zips:
                            vendor = ('글로브' if '글로브' in z_file.name else ('빌리브' if '빌리브' in z_file.name else '기타'))

                            with zipfile.ZipFile(z_file, 'r') as z:
                                g3_files = [f for f in z.namelist() if '가평3' in f and not f.endswith('/')]
                                if not g3_files: continue

                                subfolders = {}
                                for f in g3_files:
                                    parts = f.split('/')
                                    try:
                                        g_idx = next(i for i, p in enumerate(parts) if '가평3' in p)
                                        if g_idx + 1 < len(parts) - 1:
                                            subfolder_name = parts[g_idx + 1]
                                            if 'VFR3' in subfolder_name: continue
                                            subfolder_path = '/'.join(parts[:g_idx+2])
                                            if subfolder_path not in subfolders:
                                                subfolders[subfolder_path] = []
                                            subfolders[subfolder_path].append(f)
                                    except StopIteration:
                                        continue

                                for sub_path in sorted(subfolders.keys()):
                                    files_in_sub = subfolders[sub_path]
                                    pdfs = [f for f in files_in_sub if f.lower().endswith('.pdf')]

                                    num_pdf = None
                                    receipt_pdf = None
                                    for pf in pdfs:
                                        filename = os.path.splitext(os.path.basename(pf))[0].strip()
                                        if re.fullmatch(r'[\d\s\-_]+', filename):
                                            num_pdf = pf
                                        elif '거래명세서' in filename:
                                            receipt_pdf = pf

                                    center_val = ""
                                    pallet_val = ""

                                    if num_pdf or receipt_pdf:
                                        if num_pdf:
                                            try:
                                                with pdfplumber.open(io.BytesIO(z.read(num_pdf))) as pdf:
                                                    for page in pdf.pages:
                                                        text = page.extract_text() or ""
                                                        text_layout = page.extract_text(layout=True) or ""
                                                        words = page.extract_words()
                                                        all_words_text = "".join(w.get('text', '') for w in words).replace(" ", "")

                                                        if not center_val:
                                                            c_match = (re.search(r'받는\s*사람\s*:\s*([A-Za-z0-9가-힣\-]+)', text) or
                                                                       re.search(r'받는\s*사람\s*:\s*([A-Za-z0-9가-힣\-]+)', text_layout) or
                                                                       re.search(r'받는사람:([A-Za-z0-9가-힣\-]+)', all_words_text))
                                                            if c_match: center_val = c_match.group(1).strip()

                                                        if not pallet_val:
                                                            p_match = (re.search(r'팔레트\s*수량\s*:\s*(\d+)', text) or
                                                                       re.search(r'팔레트\s*수량\s*:\s*(\d+)', text_layout) or
                                                                       re.search(r'팔레트수량:(\d+)', all_words_text))
                                                            if p_match: pallet_val = p_match.group(1).strip()

                                                        if center_val and pallet_val: break
                                            except Exception as e:
                                                st.error(f"읽기 에러: {e}")

                                            all_pdfs_to_merge.append((z_file, num_pdf, 'num'))

                                        if receipt_pdf:
                                            all_pdfs_to_merge.append((z_file, receipt_pdf, 'receipt'))

                                        extracted_table_data.append((vendor, center_val, pallet_val))

                        if not all_pdfs_to_merge:
                            st.info("조건에 맞는 파일 없음")
                        else:
                            merger = PdfWriter()

                            for z_file, pf, pdf_type in all_pdfs_to_merge:
                                with zipfile.ZipFile(z_file, 'r') as z:
                                    pdf_bytes = z.read(pf)
                                    temp_reader = PdfReader(io.BytesIO(pdf_bytes))
                                    tot_pages = len(temp_reader.pages)

                                    if pdf_type == 'num':
                                        if tot_pages % 2 == 0:
                                            merger.append(io.BytesIO(pdf_bytes), pages=(0, tot_pages // 2))
                                        else:
                                            half = (tot_pages + 1) // 2
                                            if tot_pages >= 1:
                                                merger.append(io.BytesIO(pdf_bytes), pages=(0, 1))
                                            if half > 2:
                                                merger.append(io.BytesIO(pdf_bytes), pages=(2, half))
                                    elif pdf_type == 'receipt':
                                        merger.append(io.BytesIO(pdf_bytes))

                            merged_pdf_buf = io.BytesIO()
                            merger.write(merged_pdf_buf)
                            merged_pdf_buf.seek(0)

                            st.success("병합 완료")

                            if extracted_table_data:
                                df = pd.DataFrame(extracted_table_data, columns=["벤더", "센터", "팔레트"])
                                st.dataframe(df, use_container_width=True)

                            st.download_button(
                                label="📥 다운로드",
                                data=merged_pdf_buf,
                                file_name=f"가평3_{date_str}.pdf",
                                mime="application/pdf",
                                type="primary",
                                use_container_width=True
                            )

                    except Exception as e:
                        st.error(f"오류: {e}")
        st.markdown("</div>", unsafe_allow_html=True)

    # ─────────────────────────────────────────────────────────────────
    # VFR3 부착물 생성
    # ─────────────────────────────────────────────────────────────────
    with col_right:
        st.markdown("### 🏷️ [2] VFR3 부착물 생성")
        st.markdown("<div style='background-color: #f0f2f5; padding: 15px; border-radius: 10px;'>", unsafe_allow_html=True)

        CENTERS   = ["대구2", "동탄1", "인천14", "이천2", "양산1", "고양1"]
        SKU_CODES = ["7391", "7392", "7715", "8834", "1959", "1960"]
        SKU_INFO  = {
            "7391": (16, 30), "7392": (12, 45), "7715": (12, 45),
            "8834": (8,  16), "1959": (8,  96), "1960": (8,  96),
        }
        NUM_BLOCKS     = 3
        ROWS_PER_BLOCK = 6

        TOTE_CENTER_XY = (1248, 742)
        TOTE_BOX_XY    = (1318, 1056)
        TOTE_UNIT_XY   = (1525, 1056)
        TOTE_CODE_XY   = (1617, 1152)
        CONT_CENTER_XY = (1248, 742)

        def get_custom_font(size):
            font_path = "NanumSquare_acEB.ttf"
            try:
                return ImageFont.truetype(font_path, size)
            except IOError:
                return ImageFont.load_default()

        def dc(draw, text, cx, cy, font):
            draw.text((cx, cy), text, font=font, fill='black', anchor='mm')

        def dc_right(draw, text, rx, y, font):
            b = draw.textbbox((0, 0), text, font=font)
            draw.text((rx - (b[2]-b[0]), y), text, font=font, fill='black')

        def generate_attachment_pdf(rows, progress_cb=None):
            from collections import OrderedDict
            center_trucks = OrderedDict()
            for center, sku_code, plt_cnt, truck_cnt in rows:
                if center not in center_trucks:
                    center_trucks[center] = 0
                center_trucks[center] = max(center_trucks[center], truck_cnt)

            total_cont = sum(center_trucks.values())
            total_tote = sum(p for _, _, p, _ in rows)
            total = total_cont + total_tote
            container_imgs, tote_imgs = [], []
            done = 0

            f_center = get_custom_font(220)
            f_num    = get_custom_font(110)
            f_code   = get_custom_font(30)

            for center, truck_cnt in center_trucks.items():
                for _ in range(truck_cnt):
                    img = Image.open("container.jpg")
                    draw = ImageDraw.Draw(img)
                    dc(draw, center, *CONT_CENTER_XY, f_center)
                    container_imgs.append(img)
                    done += 1
                    if progress_cb: progress_cb(done, total, f"컨테이너 생성중 ({done}/{total})")

            for center, sku_code, plt_cnt, truck_cnt in rows:
                for _ in range(plt_cnt):
                    box, unit = SKU_INFO[sku_code]
                    img = Image.open("tote.jpg")
                    draw = ImageDraw.Draw(img)
                    dc(draw, center,    *TOTE_CENTER_XY, f_center)
                    dc(draw, str(box),  *TOTE_BOX_XY,   f_num)
                    dc(draw, str(unit), *TOTE_UNIT_XY,  f_num)
                    dc_right(draw, sku_code, TOTE_CODE_XY[0], TOTE_CODE_XY[1], f_code)
                    tote_imgs.append(img)
                    done += 1
                    if progress_cb: progress_cb(done, total, f"토트 생성중 ({done}/{total})")

            all_imgs = container_imgs + tote_imgs
            if not all_imgs:
                raise ValueError("생성할 페이지 없음")

            first = all_imgs[0].convert('RGB')
            rest  = [i.convert('RGB') for i in all_imgs[1:]]

            pdf_buf = io.BytesIO()
            first.save(pdf_buf, format='PDF', save_all=True, append_images=rest)
            pdf_buf.seek(0)
            return pdf_buf, len(container_imgs), len(tote_imgs)

        def reset_inputs():
            for b in range(NUM_BLOCKS):
                st.session_state[f"center_{b}"] = CENTERS[0]
                st.session_state[f"truck_{b}"]  = 0
                for r in range(ROWS_PER_BLOCK):
                    st.session_state[f"plt_{b}_{r}"] = 0

        if "initialized_att" not in st.session_state:
            reset_inputs()
            st.session_state["initialized_att"] = True

        block_cols = st.columns(3)
        for b in range(NUM_BLOCKS):
            with block_cols[b]:
                st.markdown("""<div style='background-color: #FFFF00; padding: 5px; text-align: center; border: 1px solid black; font-weight: bold;'>센터 및 트럭</div>""", unsafe_allow_html=True)
                c1, c2 = st.columns(2)
                with c1:
                    st.selectbox("센터", options=CENTERS, key=f"center_{b}", label_visibility="collapsed")
                with c2:
                    st.number_input("트럭", min_value=0, step=1, key=f"truck_{b}", label_visibility="collapsed")

                st.write("")
                st.markdown("""<div style='background-color: #FFFF00; padding: 5px; text-align: center; border: 1px solid black; font-weight: bold;'>상품 &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; 팔렛트</div>""", unsafe_allow_html=True)

                for r in range(ROWS_PER_BLOCK):
                    c_sku, c_plt = st.columns([1, 1])
                    with c_sku:
                        st.markdown(f"<div style='text-align: center; padding-top: 10px; font-weight: bold;'>{SKU_CODES[r]}</div>", unsafe_allow_html=True)
                    with c_plt:
                        st.number_input("plt", min_value=0, step=1, key=f"plt_{b}_{r}", label_visibility="collapsed")

        st.write("")
        btn1, btn2 = st.columns([1, 2])
        with btn1:
            st.button("🔄 초기화", use_container_width=True, on_click=reset_inputs)
        with btn2:
            generate_clicked = st.button("📄 생성", type="primary", use_container_width=True)

        if generate_clicked:
            rows = []
            for b in range(NUM_BLOCKS):
                center = st.session_state[f"center_{b}"]
                truck  = st.session_state[f"truck_{b}"]
                for r in range(ROWS_PER_BLOCK):
                    plt_cnt = st.session_state[f"plt_{b}_{r}"]
                    if plt_cnt > 0:
                        rows.append((center, SKU_CODES[r], plt_cnt, truck))

            if not rows:
                st.warning("수량 입력 필요")
            else:
                try:
                    prog_bar = st.progress(0)
                    stat_txt = st.empty()
                    def update_prog(done, total, msg):
                        prog_bar.progress(done / total)
                        stat_txt.text(msg)

                    pdf_bytes, cont_cnt, tote_cnt = generate_attachment_pdf(rows, progress_cb=update_prog)

                    prog_bar.empty()
                    stat_txt.empty()

                    st.success(f"생성 완료 (컨테이너 {cont_cnt}, 토트 {tote_cnt})")
                    st.download_button(
                        label="📥 PDF 다운로드",
                        data=pdf_bytes,
                        file_name="부착물.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )
                except FileNotFoundError:
                    st.error("이미지 파일 없음")
                except Exception as e:
                    st.error(f"오류: {e}")

        st.markdown("</div>", unsafe_allow_html=True)


# =====================================================================
# 1창고 탭
# =====================================================================
with tab_1창고:

    sub_a, sub_f, sub_g, sub_e = st.tabs([
        "📦 [A] 쿠팡 택배송장 추출",
        "🖨️ [F] 가평1 발주서 추출",
        "✂️ [G] 라벨 4등분 & 명세서 병합",
        "🛒 [E] 쇼핑몰 주문 취합",
    ])

    # ==================================================================
    # [A] 쿠팡 택배송장 추출 (매칭사전 제외)
    # ==================================================================
    with sub_a:
        st.markdown("### 📦 쿠팡 택배송장 양식 추출")
        st.info(
            "ZIP 또는 엑셀 파일을 업로드하면 택배송장 양식으로 변환합니다.\n"
            "ZIP 파일은 내부에서 파일명에 **'택배'** 가 포함된 엑셀만 자동 인식합니다."
        )

        uploaded_a = st.file_uploader(
            "📁 엑셀 또는 ZIP 파일 업로드 (복수 선택 가능)",
            type=["xlsx", "xls", "zip"],
            accept_multiple_files=True,
            key="uploader_a"
        )

        if st.button("🚀 추출 실행", type="primary", use_container_width=True, key="run_a"):
            if not uploaded_a:
                st.warning("파일을 먼저 업로드해주세요.")
            else:
                with st.spinner("처리중..."):
                    try:
                        def get_formatted_df_a(file_obj):
                            wb = openpyxl.load_workbook(file_obj, data_only=True)
                            ws = wb.active
                            active_sheet = ws.title
                            hidden_rows = [i for i, row in ws.row_dimensions.items() if row.hidden]
                            if hasattr(file_obj, 'seek'):
                                file_obj.seek(0)
                            df = pd.read_excel(file_obj, sheet_name=active_sheet, dtype=str)
                            to_drop = [r - 2 for r in hidden_rows if r >= 2 and (r - 2) < len(df)]
                            df = df.drop(df.index[to_drop]).reset_index(drop=True)
                            try:
                                extracted = pd.DataFrame()
                                extracted['A'] = df.iloc[:, 0].fillna('').astype(str).str.strip()
                                extracted['B'] = df.iloc[:, 1].fillna('').astype(str).str.strip()
                                extracted['C'] = df.iloc[:, 2].fillna('').astype(str).str.strip()
                                extracted['D'] = ""
                                extracted['E'] = df.iloc[:, 5].fillna('').astype(str).str.strip()
                                extracted['F'] = df.iloc[:, 11].fillna('').astype(str).str.strip()
                                extracted['G'] = "1"
                                extracted['H'] = ""
                                extracted['I'] = ""
                                return extracted
                            except IndexError:
                                return pd.DataFrame()

                        all_dfs_a = []
                        for uploaded_file in uploaded_a:
                            if uploaded_file.name.lower().endswith('.zip'):
                                with zipfile.ZipFile(uploaded_file, 'r') as z:
                                    excel_files = [
                                        f for f in z.namelist()
                                        if '택배' in f and f.endswith(('.xlsx', '.xls'))
                                        and '__MACOSX' not in f
                                        and not f.split('/')[-1].startswith('._')
                                    ]
                                    for file_name in excel_files:
                                        with z.open(file_name) as f:
                                            f_df = get_formatted_df_a(io.BytesIO(f.read()))
                                            if not f_df.empty:
                                                all_dfs_a.append(f_df)
                            elif uploaded_file.name.lower().endswith(('.xlsx', '.xls')):
                                f_df = get_formatted_df_a(uploaded_file)
                                if not f_df.empty:
                                    all_dfs_a.append(f_df)

                        if not all_dfs_a:
                            st.warning("유효한 엑셀 데이터를 찾지 못했습니다.")
                        else:
                            result_df = pd.concat(all_dfs_a, ignore_index=True)
                            result_df['B'] = result_df['B'].replace(['nan', '0.0'], '')
                            result_df = result_df[
                                result_df['A'].str.strip().ne('') &
                                result_df['A'].ne('nan')
                            ]

                            display_df = result_df.copy()
                            display_df.columns = ["A:성명", "B:전화번호", "C:주소", "D:(빈칸)", "E:품목명", "F:바코드", "G:수량", "H:박스타입", "I:메모"]

                            st.success(f"✅ 추출 완료 — {len(result_df)}건")
                            st.dataframe(display_df, use_container_width=True, height=350)

                            out_buf = io.BytesIO()
                            empty_row = pd.DataFrame([[""] * len(result_df.columns)], columns=result_df.columns)
                            export_df = pd.concat([empty_row, result_df], ignore_index=True)
                            export_df.to_excel(out_buf, index=False, header=False)
                            out_buf.seek(0)
                            today = datetime.now().strftime("%Y%m%d")
                            st.download_button(
                                label="📥 엑셀 다운로드",
                                data=out_buf,
                                file_name=f"택배송장양식_쿠팡_{today}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary",
                                use_container_width=True
                            )

                    except Exception as e:
                        st.error(f"오류: {e}")

    # ==================================================================
    # [F] 가평1 발주서 추출 및 PDF 병합
    # ==================================================================
    with sub_f:
        st.markdown("### 🖨️ 가평1 발주서 추출 및 PDF 병합")
        st.info(
            "ZIP 파일을 업로드하면 **'가평1'** 폴더 내 발주서를 자동으로 추출·병합합니다.\n"
            "글로브 → 빌리브 순으로 정렬되며, 숫자 파일명 PDF의 절반 + 거래명세서를 합칩니다."
        )

        uploaded_f = st.file_uploader(
            "📁 ZIP 파일 업로드 (글로브/빌리브, 복수 선택 가능)",
            type="zip",
            accept_multiple_files=True,
            key="uploader_f"
        )

        if st.button("🚀 추출 및 병합 실행", type="primary", use_container_width=True, key="run_f"):
            if not uploaded_f:
                st.warning("ZIP 파일을 먼저 업로드해주세요.")
            else:
                with st.spinner("처리중..."):
                    try:
                        def zip_sort_key_f(file_obj):
                            n = file_obj.name
                            if '글로브' in n: return 0
                            if '빌리브' in n: return 1
                            return 2

                        sorted_zips_f = sorted(uploaded_f, key=zip_sort_key_f)
                        date_match = re.search(r'(\d{8}|\d{6}|\d{4})', sorted_zips_f[0].name)
                        date_str = date_match.group(1) if date_match else datetime.now().strftime("%Y%m%d")

                        all_pdfs_f   = []
                        table_data_f = []

                        for z_file in sorted_zips_f:
                            vendor = '글로브' if '글로브' in z_file.name else ('빌리브' if '빌리브' in z_file.name else '기타')

                            with zipfile.ZipFile(z_file, 'r') as z:
                                g1_files = [f for f in z.namelist() if '가평1' in f and not f.endswith('/')]
                                if not g1_files:
                                    continue

                                subfolders = {}
                                for f in g1_files:
                                    parts = f.split('/')
                                    try:
                                        g_idx = next(i for i, p in enumerate(parts) if '가평1' in p)
                                        if g_idx + 1 < len(parts) - 1:
                                            if 'VFR3' in parts[g_idx + 1]:
                                                continue
                                            sf_path = '/'.join(parts[:g_idx+2])
                                            subfolders.setdefault(sf_path, []).append(f)
                                    except StopIteration:
                                        continue

                                for sub_path in sorted(subfolders.keys()):
                                    pdfs = [f for f in subfolders[sub_path] if f.lower().endswith('.pdf')]
                                    num_pdf     = None
                                    receipt_pdf = None
                                    for pf in pdfs:
                                        fname = os.path.splitext(os.path.basename(pf))[0].strip()
                                        if re.fullmatch(r'[\d\s\-_]+', fname):
                                            num_pdf = pf
                                        elif '거래명세서' in fname:
                                            receipt_pdf = pf

                                    center_val = ""
                                    pallet_val = ""

                                    if num_pdf or receipt_pdf:
                                        if num_pdf:
                                            try:
                                                with pdfplumber.open(io.BytesIO(z.read(num_pdf))) as pdf:
                                                    for page in pdf.pages:
                                                        text        = page.extract_text() or ""
                                                        text_layout = page.extract_text(layout=True) or ""
                                                        words       = page.extract_words()
                                                        all_w       = "".join(w.get('text', '') for w in words).replace(" ", "")
                                                        if not center_val:
                                                            m = (re.search(r'받는\s*사람\s*:\s*([A-Za-z0-9가-힣\-]+)', text) or
                                                                 re.search(r'받는\s*사람\s*:\s*([A-Za-z0-9가-힣\-]+)', text_layout) or
                                                                 re.search(r'받는사람:([A-Za-z0-9가-힣\-]+)', all_w))
                                                            if m: center_val = m.group(1).strip()
                                                        if not pallet_val:
                                                            m = (re.search(r'팔레트\s*수량\s*:\s*(\d+)', text) or
                                                                 re.search(r'팔레트\s*수량\s*:\s*(\d+)', text_layout) or
                                                                 re.search(r'팔레트수량:(\d+)', all_w))
                                                            if m: pallet_val = m.group(1).strip()
                                                        if center_val and pallet_val:
                                                            break
                                            except Exception:
                                                pass
                                            all_pdfs_f.append((z_file, num_pdf, 'num'))

                                        if receipt_pdf:
                                            all_pdfs_f.append((z_file, receipt_pdf, 'receipt'))

                                        table_data_f.append((vendor, center_val, pallet_val))

                        if not all_pdfs_f:
                            st.warning("조건에 맞는 가평1 파일을 찾지 못했습니다.")
                        else:
                            merger = PdfWriter()
                            for z_file, pf, pdf_type in all_pdfs_f:
                                with zipfile.ZipFile(z_file, 'r') as z:
                                    pdf_bytes   = z.read(pf)
                                    temp_reader = PdfReader(io.BytesIO(pdf_bytes))
                                    tot         = len(temp_reader.pages)
                                    if pdf_type == 'num':
                                        if tot % 2 == 0:
                                            merger.append(io.BytesIO(pdf_bytes), pages=(0, tot // 2))
                                        else:
                                            half = (tot + 1) // 2
                                            if tot >= 1:
                                                merger.append(io.BytesIO(pdf_bytes), pages=(0, 1))
                                            if half > 2:
                                                merger.append(io.BytesIO(pdf_bytes), pages=(2, half))
                                    else:
                                        merger.append(io.BytesIO(pdf_bytes))

                            out_buf = io.BytesIO()
                            merger.write(out_buf)
                            out_buf.seek(0)

                            st.success(f"✅ 병합 완료 — PDF {len(all_pdfs_f)}개 처리")

                            if table_data_f:
                                df_f = pd.DataFrame(table_data_f, columns=["벤더", "센터", "팔레트"])
                                st.dataframe(df_f, use_container_width=True, height=300)

                            st.download_button(
                                label="📥 병합 PDF 다운로드",
                                data=out_buf,
                                file_name=f"가평1_{date_str}.pdf",
                                mime="application/pdf",
                                type="primary",
                                use_container_width=True
                            )

                    except Exception as e:
                        st.error(f"오류: {e}")

    # ==================================================================
    # [G] 라벨 4등분 & 명세서 연속 병합
    # ==================================================================
    with sub_g:
        st.markdown("### ✂️ 라벨 4등분 & 명세서 연속 병합")
        st.info(
            "ZIP 파일을 업로드하면 내부 PDF를 자동으로 분류·처리합니다.\n"
            "- **Label** 파일 → 4장씩 한 페이지에 모아찍기 (4등분)\n"
            "- **ManiFest** 파일 → 원본 그대로 이어 붙이기"
        )

        uploaded_g = st.file_uploader(
            "📁 ZIP 파일 업로드 (글로브/빌리브, 복수 선택 가능)",
            type="zip",
            accept_multiple_files=True,
            key="uploader_g"
        )

        if st.button("🚀 4등분 & 병합 실행", type="primary", use_container_width=True, key="run_g"):
            if not uploaded_g:
                st.warning("ZIP 파일을 먼저 업로드해주세요.")
            else:
                with st.spinner("처리중..."):
                    try:
                        from pypdf import PageObject, Transformation

                        def zip_sort_key_g(file_obj):
                            n = file_obj.name
                            if '글로브' in n: return 0
                            if '빌리브' in n: return 1
                            return 2

                        sorted_zips_g = sorted(uploaded_g, key=zip_sort_key_g)

                        label_pages   = []
                        manifest_data = []

                        for z_file in sorted_zips_g:
                            with zipfile.ZipFile(z_file, 'r') as z:
                                for f in z.namelist():
                                    if f.endswith('/') or not f.lower().endswith('.pdf'):
                                        continue
                                    base_n    = os.path.basename(f).lower()
                                    pdf_bytes = z.read(f)
                                    if 'label' in base_n:
                                        for page in PdfReader(io.BytesIO(pdf_bytes)).pages:
                                            label_pages.append(page)
                                    elif 'manifest' in base_n:
                                        manifest_data.append(pdf_bytes)

                        if not label_pages and not manifest_data:
                            st.warning("'Label' 또는 'ManiFest' 파일을 찾지 못했습니다.")
                        else:
                            final_writer = PdfWriter()

                            # Label → 4등분
                            if label_pages:
                                w = float(label_pages[0].mediabox.width)
                                h = float(label_pages[0].mediabox.height)
                                positions = [(0, h/2), (w/2, h/2), (0, 0), (w/2, 0)]

                                for i in range(0, len(label_pages), 4):
                                    chunk = label_pages[i:i+4]
                                    blank = PageObject.create_blank_page(width=w, height=h)
                                    for j, page in enumerate(chunk):
                                        tx, ty = positions[j]
                                        try:
                                            op = Transformation().scale(sx=0.5, sy=0.5).translate(tx=tx, ty=ty)
                                            page.add_transformation(op)
                                            blank.merge_page(page)
                                        except Exception:
                                            try:
                                                blank.mergeScaledTranslatedPage(page, 0.5, tx, ty)
                                            except Exception:
                                                pass
                                    final_writer.add_page(blank)

                            # ManiFest → 이어 붙이기
                            for pdf_bytes in manifest_data:
                                for page in PdfReader(io.BytesIO(pdf_bytes)).pages:
                                    final_writer.add_page(page)

                            out_buf = io.BytesIO()
                            final_writer.write(out_buf)
                            out_buf.seek(0)

                            label_page_count = (len(label_pages) + 3) // 4
                            date_str         = datetime.now().strftime("%Y%m%d_%H%M%S")

                            st.success(
                                f"✅ 완료 — "
                                f"Label {len(label_pages)}장 → {label_page_count}페이지, "
                                f"ManiFest {len(manifest_data)}개 병합"
                            )
                            st.download_button(
                                label="📥 병합 PDF 다운로드",
                                data=out_buf,
                                file_name=f"라벨_명세서_통합_{date_str}.pdf",
                                mime="application/pdf",
                                type="primary",
                                use_container_width=True
                            )

                    except Exception as e:
                        st.error(f"오류: {e}")


    # ==================================================================
    # [E] 쇼핑몰 주문 취합
    # ==================================================================
    with sub_e:
        import json as _json_e
        import urllib.request as _urllib_req
        import urllib.parse as _urllib_parse

        # ── 기본 매핑 사전 ────────────────────────────────────────────
        DEFAULT_SHOP_MAPPING_E = {
            "누누마켓":   {"keywords":["누누마켓"],"header_row":0,"sheet":"first","skip_hidden":False,"shopname":"누누마켓(다오)","cols":{"쇼핑몰주문번호":"auto","주문자명":"받는분성명","주문자휴대폰번호":"받는분전화번호","수령자명":"받는분성명","수령자휴대폰번호":"받는분전화번호","우편번호":"[API]","주소":"받는분주소(전체, 분할)","배송메세지":"배송메세지1","온라인 상품명":"품목명","옵션명":"옵션명","주문수량":"수량","건별출고수량":"1","금액":"0"}},
            "어바웃펫":   {"keywords":["어바웃펫"],"header_row":0,"sheet":"붙여넣기시트","skip_hidden":False,"shopname":"주식회사 어바웃펫","cols":{"쇼핑몰주문번호":"주문번호","주문자명":"수취인 명","주문자휴대폰번호":"휴대폰","수령자명":"수취인 명","수령자휴대폰번호":"휴대폰","우편번호":"우편번호","주소":"도로 주소& \" \"&도로 상세 주소","배송메세지":"배송 메모","온라인 상품명":"상품 명","옵션명":"[공란]","주문수량":"배송수량","건별출고수량":"1","금액":"총 결제 금액"}},
            "롯데멤버스": {"keywords":["deliverymnglist","롯데멤버스"],"header_row":0,"sheet":"first","skip_hidden":False,"shopname":"롯데멤버스","cols":{"쇼핑몰주문번호":"주문번호","주문자명":"수취인명","주문자휴대폰번호":"연락처","수령자명":"수취인명","수령자휴대폰번호":"연락처","우편번호":"배송지우편번호","주소":"배송지","배송메세지":"배송요구사항","온라인 상품명":"상품명","옵션명":"옵션명","주문수량":"수량","건별출고수량":"1","금액":"주문금액"}},
            "펫프렌즈":   {"keywords":["다운로드_배송관리_판매자","펫프렌즈"],"header_row":0,"sheet":"first","skip_hidden":False,"shopname":"주식회사 펫프렌즈","cols":{"쇼핑몰주문번호":"주문번호","주문자명":"수령자명","주문자휴대폰번호":"수령자 연락처","수령자명":"수령자명","수령자휴대폰번호":"수령자 연락처","우편번호":"우편번호","주소":"주소& \" \"&상세주소","배송메세지":"배송메모& \" \"&공동현관 출입방법","온라인 상품명":"상품명","옵션명":"옵션값","주문수량":"수량","건별출고수량":"1","금액":"결제가"}},
            "삼성카드":   {"keywords":["삼성카드"],"header_row":0,"sheet":"first","skip_hidden":False,"shopname":"삼성카드 주식회사","cols":{"쇼핑몰주문번호":"주문번호","주문자명":"수취인","주문자휴대폰번호":"휴대폰번호","수령자명":"수취인","수령자휴대폰번호":"휴대폰번호","우편번호":"우편번호","주소":"주소","배송메세지":"고객배송요청사항","온라인 상품명":"상품명","옵션명":"[공란]","주문수량":"수량","건별출고수량":"1","금액":"공급금액"}},
            "더진엠씨":   {"keywords":["더진엠씨"],"header_row":0,"sheet":"first","skip_hidden":False,"shopname":"주식회사 더진엠씨","cols":{"쇼핑몰주문번호":"주문번호","주문자명":"수령인명","주문자휴대폰번호":"수령인연락처","수령자명":"수령인명","수령자휴대폰번호":"수령인연락처","우편번호":"우편번호","주소":"주소","배송메세지":"배송시 요청사항","온라인 상품명":"판매사상품명","옵션명":"판매사옵션명","주문수량":"주문수량","건별출고수량":"1","금액":"0"}},
            "이지앤웰스": {"keywords":["발주서_글로브","이지앤웰스"],"header_row":0,"sheet":"first","skip_hidden":False,"shopname":"이지앤웰스","cols":{"쇼핑몰주문번호":"발주번호","주문자명":"수령자","주문자휴대폰번호":"수령자휴대폰","수령자명":"수령자","수령자휴대폰번호":"수령자휴대폰","우편번호":"우편번호","주소":"주소","배송메세지":"배송메모","온라인 상품명":"상품명","옵션명":"옵션","주문수량":"수량","건별출고수량":"1","금액":"매입가"}},
            "그린마켓":   {"keywords":["그린마켓"],"header_row":0,"sheet":"first","skip_hidden":False,"shopname":"주식회사 그린마켓","cols":{"쇼핑몰주문번호":"주문번호","주문자명":"수취인명","주문자휴대폰번호":"전화번호","수령자명":"수취인명","수령자휴대폰번호":"전화번호","우편번호":"[API]","주소":"주소","배송메세지":"배송메시지","온라인 상품명":"상품명","옵션명":"[공란]","주문수량":"실수량","건별출고수량":"1","금액":"0"}},
            "블루엘엔씨": {"keywords":["블루엘엔씨"],"header_row":3,"sheet":"first","skip_hidden":False,"shopname":"주식회사 블루엘엔씨","cols":{"쇼핑몰주문번호":"auto","주문자명":"수령인","주문자휴대폰번호":"받는분전화번호","수령자명":"수령인","수령자휴대폰번호":"받는분전화번호","우편번호":"우편번호","주소":"받는분주소(전체, 분할)","배송메세지":"배송메세지1","온라인 상품명":"품목명","옵션명":"옵션","주문수량":"수량","건별출고수량":"1","금액":"0"}},
            "포스라":     {"keywords":["포스라"],"header_row":0,"sheet":"first","skip_hidden":False,"shopname":"(주)포스라","cols":{"쇼핑몰주문번호":"주문번호","주문자명":"수취인","주문자휴대폰번호":"수취인휴대폰번호","수령자명":"수취인","수령자휴대폰번호":"수취인휴대폰번호","우편번호":"우편번호","주소":"주소& \" \"&상세주소","배송메세지":"배송요청사항","온라인 상품명":"상품명","옵션명":"상품옵션","주문수량":"구매수량","건별출고수량":"1","금액":"매입가"}},
            "케이앤씨":   {"keywords":["케이앤씨"],"header_row":0,"sheet":"first","skip_hidden":False,"shopname":"케이앤씨플러스","cols":{"쇼핑몰주문번호":"auto","주문자명":"수령자명","주문자휴대폰번호":"무선전화번호","수령자명":"수령자명","수령자휴대폰번호":"무선전화번호","우편번호":"우편번호","주소":"주소","배송메세지":"유의사항","온라인 상품명":"상품명","옵션명":"[공란]","주문수량":"수량","건별출고수량":"1","금액":"0"}},
            "휘파람":     {"keywords":["휘파람"],"header_row":0,"sheet":"first","skip_hidden":False,"shopname":"(주) 휘파람","cols":{"쇼핑몰주문번호":"주문번호","주문자명":"주문자","주문자휴대폰번호":"주문자휴대폰","수령자명":"수령인","수령자휴대폰번호":"수령인휴대폰","우편번호":"배송지우편번호","주소":"배송지주소-도로명","배송메세지":"배송시문구","온라인 상품명":"대표상품명","옵션명":"옵션1","주문수량":"수량","건별출고수량":"1","금액":"금액"}},
            "리얼마인":   {"keywords":["리얼마인"],"header_row":0,"sheet":"first","skip_hidden":False,"shopname":"리얼마인","cols":{"쇼핑몰주문번호":"auto","주문자명":"받는분성명","주문자휴대폰번호":"받는분전화번호","수령자명":"받는분성명","수령자휴대폰번호":"받는분전화번호","우편번호":"[API]","주소":"받는분주소(전체, 분할)","배송메세지":"배송메세지1","온라인 상품명":"품목명","옵션명":"옵션명","주문수량":"수량","건별출고수량":"1","금액":"0"}},
            "라온커머스": {"keywords":["라온커머스"],"header_row":3,"sheet":"first","skip_hidden":True,"shopname":"라온커머스[글로브]","cols":{"쇼핑몰주문번호":"주문번호","주문자명":"이름","주문자휴대폰번호":"전화번호1","수령자명":"이름","수령자휴대폰번호":"전화번호1","우편번호":"[API]","주소":"주소","배송메세지":"배송메세지","온라인 상품명":"제품명","옵션명":"모델명","주문수량":"수량","건별출고수량":"1","금액":"단가"}},
            "서브원":     {"keywords":["ordacpcurst","서브원"],"header_row":0,"sheet":"first","skip_hidden":False,"shopname":"서브원","cols":{"쇼핑몰주문번호":"발주번호","주문자명":"수령인","주문자휴대폰번호":"수령인휴대폰번호","수령자명":"수령인","수령자휴대폰번호":"수령인휴대폰번호","우편번호":"수령인우편번호","주소":"수령인주소","배송메세지":"[공란]","온라인 상품명":"규격","옵션명":"SSP추가규격","주문수량":"수량","건별출고수량":"1","금액":"총액"}},
            "뉴퍼마켓":   {"keywords":["뉴퍼마켓"],"header_row":0,"sheet":"first","skip_hidden":False,"shopname":"뉴퍼마켓","cols":{"쇼핑몰주문번호":"주문코드","주문자명":"수령인이름","주문자휴대폰번호":"수령인연락처","수령자명":"수령인이름","수령자휴대폰번호":"수령인연락처","우편번호":"우편번호","주소":"주소","배송메세지":"배송(설치)메세지","온라인 상품명":"재고상품명","옵션명":"모델명","주문수량":"수량","건별출고수량":"1","금액":"원가(공급가)"}},
        }

        DEFAULT_INVOICE_COL_MAP_E = {
            "누누마켓":   {"col_name":"송장번호",    "header_row":0,"skip_hidden":False},
            "어바웃펫":   {"col_name":"송장번호",    "header_row":0,"skip_hidden":False},
            "롯데멤버스": {"col_name":"송장번호",    "header_row":0,"skip_hidden":False},
            "펫프렌즈":   {"col_name":"운송장번호",  "header_row":0,"skip_hidden":False},
            "삼성카드":   {"col_name":"단품명",      "header_row":0,"skip_hidden":False},
            "더진엠씨":   {"col_name":"배송번호",    "header_row":0,"skip_hidden":False},
            "이지앤웰스": {"col_name":"운송장",      "header_row":0,"skip_hidden":False},
            "그린마켓":   {"col_name":"송장번호",    "header_row":0,"skip_hidden":False},
            "블루엘엔씨": {"col_name":"운송장번호",  "header_row":3,"skip_hidden":False},
            "포스라":     {"col_name":"송장번호",    "header_row":0,"skip_hidden":False},
            "케이앤씨":   {"col_name":"운송자번호",  "header_row":0,"skip_hidden":False},
            "휘파람":     {"col_name":"송장번호",    "header_row":0,"skip_hidden":False},
            "리얼마인":   {"col_name":None,          "header_row":0,"skip_hidden":False},
            "라온커머스": {"col_name":"송장번호",    "header_row":3,"skip_hidden":True},
            "서브원":     {"col_name":None,          "header_row":0,"skip_hidden":False},
            "뉴퍼마켓":   {"col_name":"출고송장번호","header_row":0,"skip_hidden":False},
        }

        OUTPUT_COLUMNS_E = [
            "쇼핑몰주문번호","주문자명","주문자ID","주문자휴대폰번호","주문자전화번호",
            "수령자명","수령자영문명","수령자휴대폰번호","수령자전화번호","우편번호","주소",
            "배송메세지","쇼핑몰 상품코드","판매자 관리코드","온라인 상품명","옵션명",
            "주문수량","SKU코드(세트코드)","건별출고수량","배송처 코드","금액","공급가",
            "원가","실결제금액","할인금액","추가구매 옵션명","추가구매 SKU코드(세트코드)",
            "추가구매 건별출고수량","추가구매 주문수량","배송방법","배송비","사은품",
            "개인통관번호","주문일","결제완료일","국가코드","결제통화",
        ]
        ACTIVE_COLS_E = {
            "쇼핑몰주문번호","주문자명","주문자휴대폰번호","수령자명","수령자휴대폰번호",
            "우편번호","주소","배송메세지","온라인 상품명","옵션명","주문수량","건별출고수량","금액",
        }
        JUSO_KEY_E = "U01TX0FVVEgyMDI2MDMyMDExMjI0NzExNzc2MzI="

        if "e_shop_mapping" not in st.session_state:
            st.session_state["e_shop_mapping"] = _json_e.loads(_json_e.dumps(DEFAULT_SHOP_MAPPING_E))
        if "e_invoice_col_map" not in st.session_state:
            st.session_state["e_invoice_col_map"] = _json_e.loads(_json_e.dumps(DEFAULT_INVOICE_COL_MAP_E))

        # ── 헬퍼 함수 ─────────────────────────────────────────────────
        def e_identify_shop(filename, mapping):
            name_lower = os.path.splitext(os.path.basename(filename))[0].lower()
            for shop_key, info in mapping.items():
                for kw in info["keywords"]:
                    if kw.lower() in name_lower:
                        return shop_key, info
            return None, None

        def e_get_first_sheet(file_obj):
            try:
                file_obj.seek(0)
                wb = openpyxl.load_workbook(file_obj, read_only=True)
                name = wb.active.title
                wb.close()
                return name
            except Exception:
                return 0

        def e_get_hidden_indices(file_obj, sheet_name, header_row):
            try:
                file_obj.seek(0)
                wb = openpyxl.load_workbook(file_obj, data_only=True)
                ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
                data_start = header_row + 2
                hidden = set()
                for excel_row in range(data_start, (ws.max_row or 0) + 1):
                    rd = ws.row_dimensions.get(excel_row)
                    if rd and rd.hidden:
                        hidden.add(excel_row - data_start)
                wb.close()
                return hidden
            except Exception:
                return set()

        def e_clean_addr(address):
            addr = re.sub(r'^[\(\[]\s*\d{5}\s*[\)\]]\s*', '', address.strip())
            return re.sub(r'\s+', ' ', addr).strip()

        def e_extract_zip(address):
            m = re.search(r'[\(\[]\s*(\d{5})\s*[\)\]]', address)
            return m.group(1) if m else ""

        def e_call_juso(keyword):
            params = _urllib_parse.urlencode({"currentPage":1,"countPerPage":1,"keyword":keyword.strip(),"confmKey":JUSO_KEY_E,"resultType":"json"})
            url = "https://www.juso.go.kr/addrlink/addrLinkApi.do?" + params
            with _urllib_req.urlopen(url, timeout=5) as res:
                data = _json_e.loads(res.read().decode("utf-8"))
            juso = data.get("results",{}).get("juso",[])
            return juso[0].get("zipNo","") if juso else ""

        def e_get_zip(address, logs):
            if not address or not address.strip(): return ""
            z = e_extract_zip(address)
            if z: return z
            def vers(addr):
                v = [addr.strip()]
                v2 = re.sub(r'\s+', ' ', re.sub(r'\(.*?\)', '', addr)).strip()
                if v2 != v[-1]: v.append(v2)
                v3 = v2.split(',')[0].strip()
                if v3 != v[-1]: v.append(v3)
                m2 = re.match(r'^(.+?(?:\d+(?:-\d+)?)?(?:번길|로|길|대로)\s*\d+(?:-\d+)?)', v3)
                if m2:
                    v4 = m2.group(1).strip()
                    if v4 != v[-1]: v.append(v4)
                return v
            try:
                for ver in vers(address):
                    try:
                        z = e_call_juso(ver)
                        if z: return z
                    except Exception: continue
                logs.append(f"   ⚠️ 우편번호 미매칭: {address[:40]}...")
                return ""
            except Exception as ex:
                logs.append(f"   ⚠️ API 오류: {ex}")
                return ""

        def e_resolve(row_dict, df_cols, spec):
            spec = spec.strip()
            if spec == "auto": return "auto"
            if spec == "[API]": return "__API__"
            if spec == "[공란]": return ""
            if spec in ("0","1"): return spec
            if "&" in spec:
                parts = re.split(r'&', spec)
                result = []
                for p in parts:
                    p = p.strip().strip('"')
                    if p == " ": result.append(" ")
                    elif p in df_cols: result.append(str(row_dict.get(p,"")).strip())
                    else: result.append(p)
                return "".join(result).strip()
            if spec in df_cols:
                val = row_dict.get(spec, "")
                return "" if pd.isna(val) else str(val).strip()
            return ""

        def e_process_file(file_obj, filename, mapping, logs):
            shop_key, info = e_identify_shop(filename, mapping)
            if info is None:
                logs.append(f"⚠️  인식 불가: {filename}")
                return None
            shopname    = info["shopname"]
            header_row  = info["header_row"]
            sheet_spec  = info["sheet"]
            skip_hidden = info["skip_hidden"]
            col_map     = info["cols"]
            try:
                file_obj.seek(0)
                actual_sheet = e_get_first_sheet(file_obj) if sheet_spec == "first" else sheet_spec
                hidden_indices = set()
                if skip_hidden:
                    hidden_indices = e_get_hidden_indices(file_obj, actual_sheet, header_row)
                file_obj.seek(0)
                df = pd.read_excel(file_obj, sheet_name=actual_sheet, header=header_row, dtype=str).fillna("")
                if skip_hidden and hidden_indices:
                    keep = [i for i in range(len(df)) if i not in hidden_indices]
                    df = df.iloc[keep].reset_index(drop=True)
                addr_src = col_map.get("주소","")
                if addr_src and "&" not in addr_src and addr_src in df.columns:
                    df = df[df[addr_src].str.strip() != ""].reset_index(drop=True)
                elif addr_src and "&" in addr_src:
                    first_col = re.split(r'&', addr_src)[0].strip().strip('"')
                    if first_col in df.columns:
                        df = df[df[first_col].str.strip() != ""].reset_index(drop=True)
                needs_api = any(v == "[API]" for v in col_map.values())
                addr_spec = col_map.get("주소","")
                df_cols = list(df.columns)
                out_rows = []
                api_fail = 0
                for _, row in df.iterrows():
                    row_dict = row.to_dict()
                    address_val = ""
                    if needs_api and addr_spec:
                        address_val = e_resolve(row_dict, df_cols, addr_spec)
                    out_row = []
                    for col in OUTPUT_COLUMNS_E:
                        if col not in ACTIVE_COLS_E:
                            out_row.append("")
                        elif col == "주문자명":
                            out_row.append(shopname)
                        elif col == "온라인 상품명":
                            val = e_resolve(row_dict, df_cols, col_map[col]) if col in col_map else ""
                            out_row.append(val.replace("*","x"))
                        elif col == "주소":
                            val = e_resolve(row_dict, df_cols, col_map[col]) if col in col_map else ""
                            out_row.append(e_clean_addr(val))
                        elif col in col_map:
                            val = e_resolve(row_dict, df_cols, col_map[col])
                            if val == "__API__":
                                zc = e_get_zip(address_val, logs)
                                if not zc: api_fail += 1
                                out_row.append(zc)
                            else:
                                out_row.append(val)
                        else:
                            out_row.append("")
                    out_rows.append(out_row)
                result = pd.DataFrame(out_rows, columns=OUTPUT_COLUMNS_E)
                msg = f"✅ {filename}  →  {shopname}  ({len(result)}건)"
                if api_fail: msg += f"  ⚠️ 우편번호 미매칭 {api_fail}건"
                logs.append(msg)
                return result
            except Exception as ex:
                logs.append(f"❌ 오류 [{filename}]: {ex}")
                return None

        def e_write_invoice(file_obj, filename, shop_key, invoice_map, inv_cfg, shop_mapping, logs):
            if shop_key not in inv_cfg:
                logs.append(f"⚠️  [{filename}] 송장 열 설정 없음")
                return None
            info = inv_cfg[shop_key]
            col_name    = info["col_name"]
            header_row  = info["header_row"]
            skip_hidden = info["skip_hidden"]
            try:
                file_obj.seek(0)
                wb = openpyxl.load_workbook(file_obj)
                ws = wb.active
                excel_hr = header_row + 1
                header_cells = list(ws.iter_rows(min_row=excel_hr, max_row=excel_hr))[0]
                name_col_idx = None
                invoice_col_idx = None
                last_header_col = 0
                for cell in header_cells:
                    if cell.value is not None:
                        last_header_col = cell.column
                name_src = shop_mapping.get(shop_key, {}).get("cols",{}).get("수령자명","")
                for cell in header_cells:
                    if cell.value is None: continue
                    val = str(cell.value).strip()
                    if name_col_idx is None:
                        if name_src and name_src == val:
                            name_col_idx = cell.column
                        if name_col_idx is None and val in ("받는분성명","수취인명","수취인","수령인명","수령자명","수령자","이름","수령인"):
                            name_col_idx = cell.column
                    if col_name and val == col_name:
                        invoice_col_idx = cell.column
                if invoice_col_idx is None:
                    invoice_col_idx = last_header_col + 1
                    ws.cell(row=excel_hr, column=invoice_col_idx).value = "송장번호"
                    logs.append(f"   ℹ️  송장 열 미발견 → 신규 기재")
                if name_col_idx is None:
                    logs.append(f"⚠️  [{filename}] 수령자명 열을 찾을 수 없음")
                    return None
                hidden_rows = set()
                if skip_hidden:
                    for rnum, rd in ws.row_dimensions.items():
                        if rd.hidden: hidden_rows.add(rnum)
                written = 0
                not_found = 0
                for row in ws.iter_rows(min_row=excel_hr+1):
                    rnum = row[0].row
                    if rnum in hidden_rows: continue
                    name_cell = ws.cell(row=rnum, column=name_col_idx)
                    if name_cell.value is None: continue
                    name_val = str(name_cell.value).strip()
                    if name_val in invoice_map:
                        ws.cell(row=rnum, column=invoice_col_idx).value = invoice_map[name_val]
                        written += 1
                    else:
                        not_found += 1
                out_buf = io.BytesIO()
                wb.save(out_buf)
                out_buf.seek(0)
                logs.append(f"✅ [{filename}] 송장 {written}건 기재 완료")
                if not_found: logs.append(f"   ⚠️ C파일에 없는 이름 {not_found}건 (미기재)")
                return out_buf, f"{os.path.splitext(filename)[0]}_회신.xlsx"
            except Exception as ex:
                logs.append(f"❌ [{filename}] 송장 기재 오류: {ex}")
                return None

        # ── UI ────────────────────────────────────────────────────────
        st.markdown("### 🛒 쇼핑몰 주문 취합")
        main_tab_e, settings_tab_e = st.tabs(["▶ 실행", "⚙️ 매핑 설정"])

        with main_tab_e:
            col_le, col_re = st.columns(2)
            with col_le:
                st.markdown("#### 📂 A파일 — 쇼핑몰 주문 파일")
                uploaded_a_files = st.file_uploader(
                    "엑셀 파일 업로드 (복수 선택 가능)",
                    type=["xlsx","xls"], accept_multiple_files=True, key="e_a_files"
                )
                if uploaded_a_files:
                    for uf in uploaded_a_files:
                        sk, _ = e_identify_shop(uf.name, st.session_state["e_shop_mapping"])
                        badge = f"✅ {sk}" if sk else "❓ 인식불가"
                        st.markdown(f"- `{uf.name}` &nbsp; **{badge}**", unsafe_allow_html=True)
            with col_re:
                st.markdown("#### 📄 C파일 — 송장번호 파일")
                uploaded_c_file = st.file_uploader(
                    "A열:쇼핑몰명 / B열:수령자명 / C열:운송장번호",
                    type=["xlsx","xls"], accept_multiple_files=False, key="e_c_file"
                )
                if uploaded_c_file:
                    st.success(f"선택됨: `{uploaded_c_file.name}`")

            st.divider()
            btn_c1, btn_c2 = st.columns(2)
            with btn_c1:
                run_agg_e = st.button("▶ 취합 실행 → 다운로드", type="primary", use_container_width=True, key="e_run_agg")
            with btn_c2:
                run_inv_e = st.button("📦 송장번호 기재 → 다운로드", use_container_width=True, key="e_run_inv")

            log_placeholder = st.empty()

            if run_agg_e:
                if not uploaded_a_files:
                    st.warning("A파일을 먼저 업로드해주세요.")
                else:
                    with st.spinner("처리중..."):
                        logs = [f"▶ 취합 시작 — {len(uploaded_a_files)}개 파일\n"]
                        all_dfs_e = []
                        for uf in uploaded_a_files:
                            df_e = e_process_file(uf, uf.name, st.session_state["e_shop_mapping"], logs)
                            if df_e is not None and not df_e.empty:
                                all_dfs_e.append(df_e)
                        log_placeholder.code("\n".join(logs), language=None)
                        if not all_dfs_e:
                            st.error("처리된 데이터가 없습니다. 파일명을 확인해주세요.")
                        else:
                            merged_e = pd.concat(all_dfs_e, ignore_index=True)
                            if "금액" in merged_e.columns:
                                def _round_amt(v):
                                    try: return str(round(float(v))) if v != "" else ""
                                    except: return v
                                merged_e["금액"] = merged_e["금액"].apply(_round_amt)
                            tmp_buf = io.BytesIO()
                            merged_e.to_excel(tmp_buf, index=False)
                            tmp_buf.seek(0)
                            wb_e = openpyxl.load_workbook(tmp_buf)
                            ws_e = wb_e.active
                            for col_cells in ws_e.iter_cols():
                                cl = col_cells[0].column_letter
                                ci = col_cells[0].column
                                cn = OUTPUT_COLUMNS_E[ci-1] if ci-1 < len(OUTPUT_COLUMNS_E) else None
                                if cn not in ACTIVE_COLS_E:
                                    ws_e.column_dimensions[cl].hidden = True
                                else:
                                    mx = max((len(str(c.value)) if c.value else 0) for c in col_cells)
                                    ws_e.column_dimensions[cl].width = min(mx+2, 60)
                            final_buf_e = io.BytesIO()
                            wb_e.save(final_buf_e)
                            final_buf_e.seek(0)
                            today_e = datetime.now().strftime("%Y%m%d_%H%M%S")
                            st.success(f"✅ 취합 완료 — 총 {len(merged_e)}건")
                            st.download_button(
                                "📥 취합 엑셀 다운로드", data=final_buf_e,
                                file_name=f"취합_{today_e}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary", use_container_width=True
                            )

            if run_inv_e:
                if not uploaded_a_files:
                    st.warning("A파일을 먼저 업로드해주세요.")
                elif not uploaded_c_file:
                    st.warning("C파일(송장번호 파일)을 선택해주세요.")
                else:
                    with st.spinner("처리중..."):
                        logs = [f"▶ 송장번호 기재 시작 — {len(uploaded_a_files)}개 A파일\n"]
                        try:
                            uploaded_c_file.seek(0)
                            df_c_e = pd.read_excel(uploaded_c_file, header=0, dtype=str).fillna("")
                            cols_c = df_c_e.columns.tolist()
                            shop_invoice_e = {}
                            for _, row_c in df_c_e.iterrows():
                                shop = str(row_c[cols_c[0]]).strip()
                                name = str(row_c[cols_c[1]]).strip()
                                inv  = str(row_c[cols_c[2]]).strip()
                                if not shop or not name or not inv: continue
                                shop_invoice_e.setdefault(shop, {})[name] = inv

                            def _find_c_shop(shopname):
                                if shopname in shop_invoice_e: return shopname
                                for cs in shop_invoice_e:
                                    if cs in shopname or shopname in cs: return cs
                                def _core(s):
                                    return re.sub(r'(주식회사|\(주\)|\[.*?\]|\(.*?\))', '', s).strip()
                                core_s = _core(shopname)
                                for cs in shop_invoice_e:
                                    if core_s and (core_s in _core(cs) or _core(cs) in core_s): return cs
                                return None

                            result_files_e = []
                            for uf in uploaded_a_files:
                                sk, sinfo = e_identify_shop(uf.name, st.session_state["e_shop_mapping"])
                                if sinfo is None:
                                    logs.append(f"⚠️ [{uf.name}] 인식 불가 — 스킵")
                                    continue
                                sname = sinfo["shopname"]
                                matched = _find_c_shop(sname)
                                if matched is None:
                                    logs.append(f"⚠️ [{uf.name}] C파일에 '{sname}' 없음 — 스킵")
                                    continue
                                inv_map_e = shop_invoice_e[matched]
                                inv_key_e = sk if sk and sk in st.session_state["e_invoice_col_map"] else list(st.session_state["e_invoice_col_map"].keys())[0]
                                res = e_write_invoice(uf, uf.name, inv_key_e, inv_map_e, st.session_state["e_invoice_col_map"], st.session_state["e_shop_mapping"], logs)
                                if res: result_files_e.append(res)

                            log_placeholder.code("\n".join(logs), language=None)

                            if len(result_files_e) == 1:
                                buf_inv, fname_inv = result_files_e[0]
                                st.success("✅ 송장번호 기재 완료")
                                st.download_button("📥 회신 파일 다운로드", data=buf_inv, file_name=fname_inv,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    type="primary", use_container_width=True)
                            elif len(result_files_e) > 1:
                                st.success(f"✅ {len(result_files_e)}개 파일 기재 완료")
                                zip_buf_e = io.BytesIO()
                                with zipfile.ZipFile(zip_buf_e, "w") as zf:
                                    for buf_inv, fname_inv in result_files_e:
                                        zf.writestr(fname_inv, buf_inv.read())
                                zip_buf_e.seek(0)
                                st.download_button("📥 전체 회신 ZIP 다운로드", data=zip_buf_e,
                                    file_name=f"송장회신_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                                    mime="application/zip", type="primary", use_container_width=True)
                            else:
                                st.warning("기재된 파일이 없습니다.")
                        except Exception as ex_inv:
                            st.error(f"오류: {ex_inv}")

        with settings_tab_e:
            st.markdown("#### ⚙️ 쇼핑몰 매핑 설정")
            st.info("쇼핑몰별 파일명 키워드, 헤더 행, 열 매핑, 거래처명을 수정할 수 있습니다. 수정 후 **저장** 버튼을 눌러야 적용됩니다.")

            cfg_t1, cfg_t2, cfg_t3 = st.tabs(["🗺️ 열 매핑 수정", "➕ 쇼핑몰 추가/삭제", "📦 송장 열 설정"])

            with cfg_t1:
                mapping_cfg = st.session_state["e_shop_mapping"]
                sel_shop = st.selectbox("수정할 쇼핑몰", list(mapping_cfg.keys()), key="cfg_sel")
                if sel_shop:
                    info_s = mapping_cfg[sel_shop]
                    c1s, c2s = st.columns(2)
                    with c1s:
                        new_sn  = st.text_input("거래처명(shopname)", value=info_s["shopname"], key=f"sn_{sel_shop}")
                        new_hr  = st.number_input("헤더 행 (0-based)", value=int(info_s["header_row"]), min_value=0, step=1, key=f"hr_{sel_shop}")
                        new_sh  = st.text_input("시트명 (first=활성)", value=str(info_s["sheet"]), key=f"sh_{sel_shop}")
                    with c2s:
                        new_sk  = st.checkbox("숨긴 행 제외", value=bool(info_s["skip_hidden"]), key=f"sk_{sel_shop}")
                        new_kw  = st.text_input("파일명 키워드 (쉼표 구분)", value=", ".join(info_s["keywords"]), key=f"kw_{sel_shop}")
                    st.markdown("**열 매핑** — 특수값: `auto` / `[API]` / `[공란]` / `0` / `1` / `열A&\" \"&열B`")
                    COL_LABELS_E = [
                        ("쇼핑몰주문번호","주문번호 열"),("주문자명","주문자명 열"),
                        ("주문자휴대폰번호","주문자 휴대폰"),("수령자명","수령자명 열"),
                        ("수령자휴대폰번호","수령자 휴대폰"),("우편번호","우편번호 ([API]=자동조회)"),
                        ("주소","주소 열"),("배송메세지","배송메세지 열"),
                        ("온라인 상품명","상품명 열"),("옵션명","옵션명 열"),
                        ("주문수량","수량 열"),("건별출고수량","건별출고수량"),("금액","금액 열 (0=공란)"),
                    ]
                    new_cols_cfg = {}
                    cp = st.columns(2)
                    for idx2, (ck, cl_lbl) in enumerate(COL_LABELS_E):
                        with cp[idx2 % 2]:
                            new_cols_cfg[ck] = st.text_input(cl_lbl, value=info_s["cols"].get(ck,""), key=f"col_{sel_shop}_{ck}")
                    if st.button("💾 저장", type="primary", key=f"save_{sel_shop}"):
                        st.session_state["e_shop_mapping"][sel_shop] = {
                            "keywords": [k.strip() for k in new_kw.split(",") if k.strip()],
                            "header_row": int(new_hr), "sheet": new_sh.strip(),
                            "skip_hidden": new_sk, "shopname": new_sn.strip(), "cols": new_cols_cfg,
                        }
                        st.success(f"✅ '{sel_shop}' 저장 완료!")
                        st.rerun()
                    if st.button("🔄 이 쇼핑몰 기본값 복원", key=f"rst_{sel_shop}"):
                        if sel_shop in DEFAULT_SHOP_MAPPING_E:
                            st.session_state["e_shop_mapping"][sel_shop] = _json_e.loads(_json_e.dumps(DEFAULT_SHOP_MAPPING_E[sel_shop]))
                            st.success("복원 완료!")
                            st.rerun()

            with cfg_t2:
                st.markdown("##### ➕ 새 쇼핑몰 추가")
                new_key = st.text_input("쇼핑몰 키 (예: 쿠팡)", key="new_shop_key")
                if st.button("추가", key="add_shop"):
                    nk = new_key.strip()
                    if not nk: st.warning("키를 입력하세요.")
                    elif nk in st.session_state["e_shop_mapping"]: st.warning("이미 존재합니다.")
                    else:
                        st.session_state["e_shop_mapping"][nk] = {
                            "keywords":[nk],"header_row":0,"sheet":"first","skip_hidden":False,"shopname":nk,
                            "cols":{"쇼핑몰주문번호":"주문번호","주문자명":"수령자명","주문자휴대폰번호":"휴대폰번호","수령자명":"수령자명","수령자휴대폰번호":"휴대폰번호","우편번호":"우편번호","주소":"주소","배송메세지":"배송메세지","온라인 상품명":"상품명","옵션명":"[공란]","주문수량":"수량","건별출고수량":"1","금액":"0"}
                        }
                        st.session_state["e_invoice_col_map"][nk] = {"col_name":"송장번호","header_row":0,"skip_hidden":False}
                        st.success(f"'{nk}' 추가 완료! [열 매핑 수정] 탭에서 설정하세요.")
                        st.rerun()
                st.markdown("---")
                st.markdown("##### 🗑️ 쇼핑몰 삭제")
                del_s = st.selectbox("삭제할 쇼핑몰", list(st.session_state["e_shop_mapping"].keys()), key="del_shop")
                if st.button("삭제", type="primary", key="del_shop_btn"):
                    del st.session_state["e_shop_mapping"][del_s]
                    if del_s in st.session_state["e_invoice_col_map"]: del st.session_state["e_invoice_col_map"][del_s]
                    st.success(f"'{del_s}' 삭제 완료!")
                    st.rerun()
                st.markdown("---")
                if st.button("🔄 전체 기본값으로 초기화", key="reset_all"):
                    st.session_state["e_shop_mapping"]    = _json_e.loads(_json_e.dumps(DEFAULT_SHOP_MAPPING_E))
                    st.session_state["e_invoice_col_map"] = _json_e.loads(_json_e.dumps(DEFAULT_INVOICE_COL_MAP_E))
                    st.success("전체 초기화 완료!")
                    st.rerun()

            with cfg_t3:
                st.markdown("##### 📦 쇼핑몰별 송장 열 설정")
                st.info("A파일에서 송장번호를 기재할 열 이름과 헤더 행을 설정합니다. col_name이 빈칸이면 마지막 열 다음에 자동 추가됩니다.")
                inv_cfg_ui = st.session_state["e_invoice_col_map"]
                inv_sel = st.selectbox("쇼핑몰 선택", list(inv_cfg_ui.keys()), key="inv_sel")
                if inv_sel:
                    cur_inv = inv_cfg_ui[inv_sel]
                    ic1, ic2, ic3 = st.columns(3)
                    with ic1:
                        new_icol = st.text_input("송장 열 이름 (없으면 비워두기)", value=cur_inv["col_name"] or "", key=f"icol_{inv_sel}")
                    with ic2:
                        new_ihr  = st.number_input("헤더 행", value=int(cur_inv["header_row"]), min_value=0, step=1, key=f"ihr_{inv_sel}")
                    with ic3:
                        new_isk  = st.checkbox("숨긴 행 제외", value=bool(cur_inv["skip_hidden"]), key=f"isk_{inv_sel}")
                    if st.button("💾 저장", type="primary", key=f"isave_{inv_sel}"):
                        st.session_state["e_invoice_col_map"][inv_sel] = {
                            "col_name": new_icol.strip() if new_icol.strip() else None,
                            "header_row": int(new_ihr), "skip_hidden": new_isk,
                        }
                        st.success(f"✅ '{inv_sel}' 송장 설정 저장 완료!")
                        st.rerun()
