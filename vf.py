import streamlit as st
import pandas as pd
import zipfile
import io
import openpyxl
import re
import base64
from io import BytesIO
from barcode import Code128
from barcode.writer import ImageWriter
from datetime import datetime, timedelta

import gspread
from google.oauth2.service_account import Credentials
import streamlit.components.v1 as components

st.set_page_config(page_title="벤더플렉스 입고 도우미", layout="wide")

def get_barcode_base64(data):
    if not data or data == "-": 
        return ""
    try:
        rv = BytesIO()
        options = {'write_text': True, 'module_width': 0.25, 'module_height': 10.0, 'font_size': 12}
        Code128(str(data), writer=ImageWriter()).write(rv, options=options)
        return f"data:image/png;base64,{base64.b64encode(rv.getvalue()).decode()}"
    except Exception as e:
        return ""

@st.cache_data(ttl=600)
def load_google_sheet():
    try:
        secret_info = st.secrets["gcp_service_account"]
        scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
        creds = Credentials.from_service_account_info(secret_info, scopes=scopes)
        client = gspread.authorize(creds)
        sheet_id = "1J5RwYs3IVCm9f0IsCjwtrSerOGdx_J3f3r0o72BgrTA"
        worksheet = client.open_by_key(sheet_id).worksheet("시트1")
        data = worksheet.get_all_values()
        if data:
            df = pd.DataFrame(data[1:], columns=data[0])
            return df
        else:
            return pd.DataFrame()
    except Exception as e:
        st.error(f"구글 시트를 불러오지 못했어요. (오류: {e})")
        return pd.DataFrame()

df_sheet = load_google_sheet()

st.markdown("""
<style>
.block-container {
    padding-top: 0rem !important;
}
header[data-testid="stHeader"] {
    display: none !important;
}
.fixed-top-bar {
    position: fixed;
    top: 0;
    left: 0;
    width: 100vw;
    background-color: white;
    z-index: 999999;
    padding: 12px 30px 14px 30px;
    border-bottom: 2px solid #ddd;
    box-shadow: 0 4px 10px rgba(0,0,0,0.1);
    box-sizing: border-box;
}
.main-content {
    margin-top: 220px;
}
.top-barcode-title {
    font-size: 20px;
    font-weight: bold;
    text-align: center;
    margin-bottom: 8px;
}
.fixed-top-bar img {
    max-width: 380px;
    max-height: 160px;
    width: 100%;
    object-fit: contain;
    display: block;
    margin: 0 auto;
}
iframe { margin-bottom: 0 !important; }
</style>
""", unsafe_allow_html=True)

workbench_img = get_barcode_base64("466-RCRT1-1-1")
ibc_placeholder_img = get_barcode_base64("IBC")

st.markdown(f"""
<div class="fixed-top-bar">
    <div style="display: flex; justify-content: space-around; align-items: center; gap: 20px;">
        <div style="text-align: center; flex: 1; min-width: 0;">
            <div class="top-barcode-title">🏷️ 토트 바코드</div>
            <img src="{workbench_img}">
        </div>
        <div style="text-align: center; flex: 1; min-width: 0;">
            <div class="top-barcode-title">🏷️ IBC 바코드</div>
            <img src="{ibc_placeholder_img}" id="ibc-barcode-img">
        </div>
        <div style="text-align: center; flex: 1; min-width: 0; display: flex; flex-direction: column; align-items: center;">
            <div class="top-barcode-title">✏️ IBC 바코드 입력</div>
            <input 
                id="ibc-input"
                type="text"
                placeholder="숫자 입력"
                style="font-size: 22px; padding: 12px; width: 80%; max-width: 300px; border: 2px solid #ccc; border-radius: 6px; text-align: center; box-sizing: border-box; display: block; margin: 0 auto;"
                oninput="updateIBC(this.value)"
            >
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

components.html("""
<script src="https://cdn.jsdelivr.net/npm/jsbarcode@3.11.5/dist/JsBarcode.all.min.js"></script>
<canvas id="ibc-canvas" style="display:none;"></canvas>
<script>
function updateIBC(val) {
    const fullCode = val ? 'IBC' + val : 'IBC';
    try {
        const canvas = document.getElementById('ibc-canvas');
        JsBarcode(canvas, fullCode, {
            format: "CODE128",
            width: 2,
            height: 80,
            displayValue: true,
            fontSize: 16
        });
        const dataUrl = canvas.toDataURL('image/png');
        window.parent.document.getElementById('ibc-barcode-img').src = dataUrl;
    } catch(e) {}
}
window.parent.document.getElementById('ibc-input').addEventListener('input', function() {
    updateIBC(this.value);
});
</script>
""", height=0)

st.markdown('<div class="main-content">', unsafe_allow_html=True)

st.markdown("<h3 style='text-align: center;'>📁 파일 업로드</h3>", unsafe_allow_html=True)
uploaded_files = st.file_uploader(
    "ZIP 파일 또는 XLSX 파일을 선택해주세요. (복수 선택 가능)",
    type=["zip", "xlsx"],
    accept_multiple_files=True
)

st.divider()

if uploaded_files:
    now_kst = (datetime.utcnow() + timedelta(hours=9)).strftime('%Y-%m-%d %H:%M:%S')
    file_names = [f.name for f in uploaded_files]
    print(f"👀 [{now_kst} KST] Gotcha! (파일명: {', '.join(file_names)})")

    # ✅ status: "not_received"(확정), "unconfirmed"(미확정), "already_received"(기입고)
    po_numbers = []
    extracted_data = []

    def process_xlsx(ws, po_num, status):
        for ri in range(1, ws.max_row + 1):
            cv = str(ws.cell(ri, 3).value or "").strip()
            if cv.startswith("PL") or cv.startswith("880"):
                qv = str(ws.cell(ri - 1, 8).value or "0").strip()
                extracted_data.append({
                    "po": po_num,
                    "barcode": cv,
                    "qty": qv,
                    "status": status
                })

    def check_status(ws):
        """
        발주서 상태 판단:
        1. Q20 셀이 '입고금액'이 아니면 → 'unconfirmed' (미확정, 빨간글씨)
        2. Q20 셀이 '입고금액'이면서:
           - H열과 I열(22행부터 2행씩 병합) 수량이 모두 일치 → 'already_received' (기입고, 회색)
           - H열과 I열 수량이 하나라도 다름              → 'not_received' (미입고, 정상출력)
        """
        # Step 1: 미확정 여부 (Q=17열, 20행)
        cell_val = str(ws.cell(20, 17).value or "").strip()
        if cell_val != "입고금액":
            return "unconfirmed"

        # Step 2: H열(8열)과 I열(9열) 수량 비교 (22행부터 2행 간격)
        all_match = True
        has_data = False
        row = 22
        while row <= ws.max_row:
            h_val = ws.cell(row, 8).value
            i_val = ws.cell(row, 9).value
            if h_val is None and i_val is None:
                break
            try:
                h_num = float(str(h_val).replace(",", "").strip() or "0")
            except:
                h_num = 0
            try:
                i_num = float(str(i_val).replace(",", "").strip() or "0")
            except:
                i_num = 0
            if h_num > 0:
                has_data = True
                if h_num != i_num:
                    all_match = False
                    break
            row += 2

        if has_data and all_match:
            return "already_received"
        else:
            return "not_received"

    for uploaded_file in uploaded_files:
        filename = uploaded_file.name

        if filename.endswith(".zip"):
            with zipfile.ZipFile(uploaded_file, 'r') as z:
                for fi in z.infolist():
                    if fi.filename.endswith('.xlsx') and not fi.filename.startswith('~'):
                        fn = fi.filename.split('/')[-1]
                        po_num = re.sub(r'[^0-9]', '', fn)
                        with z.open(fi) as f:
                            wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
                            ws = wb.active
                            status = check_status(ws)
                            if po_num and (po_num, status) not in po_numbers:
                                po_numbers.append((po_num, status))
                            process_xlsx(ws, po_num, status)

        elif filename.endswith(".xlsx"):
            po_num = re.sub(r'[^0-9]', '', filename)
            wb = openpyxl.load_workbook(io.BytesIO(uploaded_file.read()), data_only=True)
            ws = wb.active
            status = check_status(ws)
            if po_num and (po_num, status) not in po_numbers:
                po_numbers.append((po_num, status))
            process_xlsx(ws, po_num, status)

    st.markdown("<h3 style='text-align: center;'>[ 추출된 발주번호 ]</h3>", unsafe_allow_html=True)
    st.write("")
    
    po_cols = st.columns(4)
    for idx, (po, status) in enumerate(po_numbers):
        with po_cols[idx % 4]:
            if status == "not_received":
                text_color = "#222222"
                label = f"📝 {po}"
            elif status == "unconfirmed":
                text_color = "#e00000"
                label = f"📝 {po} ⚠️ 미확정"
            else:  # already_received
                text_color = "#888888"
                label = f"📝 {po} 🔒 기입고"

            st.markdown(
                f"<div style='text-align:center; font-size:18px; font-weight:bold; "
                f"margin-bottom:4px; color:{text_color};'>{label}</div>",
                unsafe_allow_html=True
            )
            img_b64 = get_barcode_base64(po)

            if status == "not_received":
                st.markdown(
                    f"<div style='text-align:center;'>"
                    f"<img src='{img_b64}' style='max-width:300px;'></div>",
                    unsafe_allow_html=True
                )
            elif status == "unconfirmed":
                st.markdown(
                    f"<div style='text-align:center; border: 2px solid #e00000; "
                    f"border-radius:6px; display:inline-block; padding:4px;'>"
                    f"<img src='{img_b64}' style='max-width:300px; "
                    f"filter: sepia(1) saturate(5) hue-rotate(-10deg);'></div>",
                    unsafe_allow_html=True
                )
            else:  # already_received - 바코드 안 찍힘
                st.markdown(
                    f"<div style='text-align:center; border: 2px solid #888888; "
                    f"border-radius:6px; display:inline-block; padding:8px; "
                    f"background:#f5f5f5; color:#888; font-size:14px;'>"
                    f"🔒 기입고 발주서<br>바코드 미출력</div>",
                    unsafe_allow_html=True
                )

    st.divider()

    has_unconfirmed = any(s == "unconfirmed" for _, s in po_numbers)
    has_already_received = any(s == "already_received" for _, s in po_numbers)

    extra_label = ""
    if has_unconfirmed:
        extra_label += " <span style='color:#e00000; font-size:16px;'>* 미확정 발주서</span>"
    if has_already_received:
        extra_label += " <span style='color:#888888; font-size:16px;'>* 기입고 발주서</span>"

    st.markdown(
        f"<h3 style='text-align: center;'>📋 상품 출력 목록{extra_label}</h3>",
        unsafe_allow_html=True
    )

    html_table = """
    <style>
    .table-container { max-height: 800px; overflow-y: auto; border: 1px solid #ddd; }
    table { width: 100%; border-collapse: collapse; text-align: center; }
    th { position: sticky; top: 0; background-color: #f4f4f4; padding: 10px;
         border-bottom: 2px solid #ccc; z-index: 10; font-size: 16px;}
    td { padding: 15px; border-bottom: 1px solid #eee; vertical-align: middle; }
    img { max-width: 280px; height: auto; }
    .product-name { font-size: 20px; text-align: left; padding-left: 10px; }
    .unconfirmed-name { font-size: 20px; text-align: left; padding-left: 10px; color: #e00000; }
    .already-received-name { font-size: 20px; text-align: left; padding-left: 10px; color: #888888; }
    </style>
    <div class="table-container">
    <table>
        <tr>
            <th>상품바코드</th>
            <th>상품명</th>
            <th>확정수량</th>
            <th>로케이션 바코드</th>
        </tr>
    """

    inventory_rows = []
    seen = set()

    for item in extracted_data:
        prod_barcode = item["barcode"]
        prod_qty = item["qty"]
        status = item["status"]
        prod_name = "매칭 실패"
        loc_num = "0"
        # ✅ 열 이름으로 창고 재고 불러오기
        w1 = w2 = w3 = w4 = "-"

        if not df_sheet.empty:
            match_row = df_sheet[df_sheet.iloc[:, 0].astype(str).str.strip() == prod_barcode]
            if not match_row.empty:
                if len(match_row.columns) > 2:
                    prod_name = str(match_row.iloc[0, 2])
                # ✅ 열 이름으로 접근 (인덱스 대신)
                w1 = str(match_row['1창고 (007)'].iloc[0]) if '1창고 (007)' in match_row.columns else "-"
                w2 = str(match_row['2창고 (012)'].iloc[0]) if '2창고 (012)' in match_row.columns else "-"
                w3 = str(match_row['3창고 (017)'].iloc[0]) if '3창고 (017)' in match_row.columns else "-"
                w4 = str(match_row['4창고 (018)'].iloc[0]) if '4창고 (018)' in match_row.columns else "-"
                if len(match_row.columns) > 11:
                    loc_num = str(match_row.iloc[0, 11])

        if status == "not_received":
            img_prod_tag = f'<img src="{get_barcode_base64(prod_barcode)}">'
            name_class = "product-name"
            qty_style = "font-size: 24px; font-weight: bold;"
        elif status == "unconfirmed":
            img_prod_tag = (
                f'<div style="color:#e00000; font-size:13px; font-weight:bold;">'
                f'⚠️ 미확정<br>{prod_barcode}</div>'
            )
            name_class = "unconfirmed-name"
            qty_style = "font-size: 24px; font-weight: bold; color: #e00000;"
        else:  # already_received - 바코드 안 찍힘
            img_prod_tag = (
                f'<div style="color:#888888; font-size:13px; font-weight:bold;">'
                f'🔒 기입고<br>{prod_barcode}</div>'
            )
            name_class = "already-received-name"
            qty_style = "font-size: 24px; font-weight: bold; color: #888888;"

        img_loc = get_barcode_base64(f"466-A1-1-{loc_num}")

        if status == "not_received":
            loc_tag = f'<img src="{img_loc}">'
        elif status == "unconfirmed":
            loc_tag = f'<img src="{img_loc}" style="filter: sepia(1) saturate(5) hue-rotate(-10deg);">'
        else:  # already_received
            loc_tag = f'<img src="{img_loc}" style="filter: grayscale(100%); opacity: 0.5;">'

        html_table += f"""
        <tr>
            <td>{img_prod_tag}</td>
            <td class="{name_class}">{prod_name}</td>
            <td style="{qty_style}">{prod_qty}</td>
            <td>{loc_tag}</td>
        </tr>
        """

        if prod_barcode not in seen:
            seen.add(prod_barcode)
            inventory_rows.append({
                "상품명": prod_name,
                "1창고": w1,
                "2창고": w2,
                "3창고": w3,
                "4창고": w4,
                "status": status
            })

    html_table += "</table></div>"

    inv_section = """
    <div style="margin-top: 24px;">
        <div style="font-size: 22px; font-weight: bold; margin-bottom: 8px; text-align: center;">📦 이카운트 재고 현황</div>
        <style>
        .inv-container { border: 1px solid #ddd; border-radius: 6px; overflow: hidden; }
        .inv-table { width: 100%; border-collapse: collapse; text-align: center; }
        .inv-table th {
            background-color: #4A90D9; color: white;
            padding: 12px 8px; font-size: 20px;
        }
        .inv-table td {
            padding: 12px 8px; border-bottom: 1px solid #eee;
            font-size: 19px; vertical-align: middle;
        }
        .inv-table tr:nth-child(even) { background-color: #f9f9f9; }
        .inv-table tr:hover { background-color: #EBF5FB; }
        .inv-name { text-align: center; font-weight: 500; }
        .inv-name-red { text-align: center; font-weight: 500; color: #e00000; }
        .inv-name-gray { text-align: center; font-weight: 500; color: #888888; }
        .inv-num  { font-weight: bold; color: #1a5276; }
        </style>
        <div class="inv-container">
        <table class="inv-table">
            <tr>
                <th>상품명</th><th>1창고</th><th>2창고</th><th>3창고</th><th>4창고</th>
            </tr>
    """

    for row in inventory_rows:
        if row["status"] == "not_received":
            name_cls = "inv-name"
        elif row["status"] == "unconfirmed":
            name_cls = "inv-name-red"
        else:
            name_cls = "inv-name-gray"

        inv_section += f"""
            <tr>
                <td class="{name_cls}">{row['상품명']}</td>
                <td class="inv-num">{row['1창고']}</td>
                <td class="inv-num">{row['2창고']}</td>
                <td class="inv-num">{row['3창고']}</td>
                <td class="inv-num">{row['4창고']}</td>
            </tr>
        """

    inv_section += "</table></div></div>"

    combined_html = html_table + inv_section

    product_height = 100 + len(extracted_data) * 180
    inv_height = 60 + len(inventory_rows) * 52
    total_height = product_height + inv_height + 100

    components.html(combined_html, height=total_height, scrolling=True)

st.markdown('</div>', unsafe_allow_html=True)
