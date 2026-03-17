import warnings
warnings.filterwarnings('ignore')

import streamlit as st
import json
import re
import zipfile
import io
import os
from datetime import datetime, timedelta

# ══════════════════════════════════════════════════════════════════
# Page Config
# ══════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="벤더플렉스 WMS",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ══════════════════════════════════════════════════════════════════
# 시스템 설정 상수
# ══════════════════════════════════════════════════════════════════
LOC_PREFIX   = "466-A1-1-"
ACTUAL_ID    = "1J5RwYs3IVCm9f0IsCjwtrSerOGdx_J3f3r0o72BgrTA"   # 서비스 계정
VENDOR_ID    = "10ZbUts1AfX7uscAGs7MAxBIiFwqp7jIgabz8nbjUOjQ"   # 구글 계정 OAuth
WAREHOUSE_ID = "1T2V7w2dM9Zcl0DydJJSb5nK0y8GvZJaVRHSH-f61ses"  # 구글 계정 OAuth

SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

# ══════════════════════════════════════════════════════════════════
# 헬퍼 함수
# ══════════════════════════════════════════════════════════════════
def _safe_int(val, default=0):
    try:
        cleaned = re.sub(r"[^\d.-]", "", str(val or ""))
        if not cleaned:
            return default
        return int(float(cleaned))
    except Exception:
        return default


def col_letter(n):
    s = ""
    n += 1
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def build_sa_service(creds_dict: dict = None):
    """서비스 계정으로 Sheets 서비스 생성.
    creds_dict가 없으면 Streamlit Secrets의 [gcp_service_account] 사용."""
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    if creds_dict is None:
        creds_dict = dict(st.secrets["gcp_service_account"])
    creds = service_account.Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    return build("sheets", "v4", credentials=creds, cache_discovery=False)


def build_oauth_service_from_token(token: str, refresh_token: str, client_id: str, client_secret: str):
    """저장된 토큰으로 Sheets 서비스 생성 + 만료 시 자동 갱신"""
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build
    creds = Credentials(
        token=token,
        refresh_token=refresh_token,
        token_uri="https://oauth2.googleapis.com/token",
        client_id=client_id,
        client_secret=client_secret,
        scopes=SCOPES,
    )
    # 액세스 토큰 만료 시 refresh_token으로 자동 갱신
    if not creds.valid:
        creds.refresh(Request())
    return build("sheets", "v4", credentials=creds, cache_discovery=False)


def get_oauth_auth_url(redirect_uri: str) -> str:
    """Secrets의 OAuth 클라이언트 정보로 인증 URL 생성"""
    from google_auth_oauthlib.flow import Flow
    client_config = {
        "web": {
            "client_id":     st.secrets["OAUTH_CLIENT_ID"],
            "client_secret": st.secrets["OAUTH_CLIENT_SECRET"],
            "auth_uri":      "https://accounts.google.com/o/oauth2/auth",
            "token_uri":     "https://oauth2.googleapis.com/token",
            "redirect_uris": [redirect_uri],
        }
    }
    flow = Flow.from_client_config(client_config, scopes=SCOPES, redirect_uri=redirect_uri)
    auth_url, state = flow.authorization_url(
        access_type="offline",
        prompt="consent",
    )
    st.session_state["_oauth_state"]        = state
    st.session_state["_oauth_redirect_uri"] = redirect_uri
    return auth_url


def exchange_oauth_code(code: str) -> dict:
    """인증 코드 → 토큰 교환. 반환값을 Secrets에 저장하도록 안내."""
    from google_auth_oauthlib.flow import Flow
    client_config = {
        "web": {
            "client_id":     st.secrets["OAUTH_CLIENT_ID"],
            "client_secret": st.secrets["OAUTH_CLIENT_SECRET"],
            "auth_uri":      "https://accounts.google.com/o/oauth2/auth",
            "token_uri":     "https://oauth2.googleapis.com/token",
            "redirect_uris": [st.session_state["_oauth_redirect_uri"]],
        }
    }
    flow = Flow.from_client_config(
        client_config, scopes=SCOPES,
        state=st.session_state.get("_oauth_state"),
        redirect_uri=st.session_state["_oauth_redirect_uri"],
    )
    flow.fetch_token(code=code)
    c = flow.credentials
    return {
        "token":         c.token,
        "refresh_token": c.refresh_token,
        "client_id":     c.client_id,
        "client_secret": c.client_secret,
    }


def _get_redirect_uri() -> str:
    """현재 Streamlit 앱 URL을 redirect_uri로 반환 (쿼리스트링 제거)"""
    try:
        raw = st.context.url
    except Exception:
        raw = "http://localhost:8501"
    return raw.split("?")[0].rstrip("/")


# ══════════════════════════════════════════════════════════════════
# Session State 초기화
# ══════════════════════════════════════════════════════════════════
def init_state():
    defaults = {
        "page":               "login",
        "svc_sa":             None,
        "svc_oauth":          None,
        "vendor_date":        datetime.now(),
        "adjust_log":         [],
        "items_data":         {},
        "_oauth_state":       None,
        "_oauth_redirect_uri": None,
        "_first_token_shown": False,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()


# ══════════════════════════════════════════════════════════════════
# 시작 시 자동 인증 (Secrets에 토큰이 있으면 즉시 복원)
# ══════════════════════════════════════════════════════════════════
def _try_auto_sa():
    """Secrets에 [gcp_service_account]가 있으면 자동으로 SA 서비스 복원"""
    if st.session_state.svc_sa is not None:
        return True
    try:
        if "gcp_service_account" not in st.secrets:
            return False
        svc = build_sa_service()
        svc.spreadsheets().get(spreadsheetId=ACTUAL_ID).execute()
        st.session_state.svc_sa = svc
        return True
    except Exception:
        return False

_sa_auto_ok = _try_auto_sa()

def _try_auto_oauth():
    """OAUTH_REFRESH_TOKEN이 Secrets에 있으면 자동으로 OAuth 서비스 복원"""
    if st.session_state.svc_oauth is not None:
        return True
    try:
        token         = st.secrets.get("OAUTH_TOKEN", "")
        refresh_token = st.secrets["OAUTH_REFRESH_TOKEN"]   # 없으면 KeyError → 최초 로그인 필요
        client_id     = st.secrets["OAUTH_CLIENT_ID"]
        client_secret = st.secrets["OAUTH_CLIENT_SECRET"]
        svc = build_oauth_service_from_token(token, refresh_token, client_id, client_secret)
        # 연결 테스트
        svc.spreadsheets().get(spreadsheetId=VENDOR_ID).execute()
        st.session_state.svc_oauth = svc
        return True
    except Exception:
        return False

_oauth_auto_ok = _try_auto_oauth()

# ── OAuth 콜백 코드 처리 (최초 1회 로그인 후 리디렉션) ─────────────
_qp = st.query_params
if "code" in _qp and st.session_state.svc_oauth is None:
    try:
        token_dict = exchange_oauth_code(_qp["code"])
        st.session_state.svc_oauth = build_oauth_service_from_token(
            token_dict["token"],
            token_dict["refresh_token"],
            token_dict["client_id"],
            token_dict["client_secret"],
        )
        st.session_state["_new_token_dict"] = token_dict  # 사용자에게 Secrets 저장 안내용
        st.session_state["_first_token_shown"] = False
        st.query_params.clear()
        if st.session_state.svc_sa is not None:
            st.session_state.page = "main"
        st.rerun()
    except Exception as e:
        st.error(f"OAuth 토큰 교환 실패: {e}")
        st.query_params.clear()


# ══════════════════════════════════════════════════════════════════
# CSS — 레트로 게임풍 디자인 (원본 색상 팔레트 그대로)
# ══════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Nanum+Square:wght@400;700;800&display=swap');

:root {
    --c-bg:        #EDD9A3;
    --c-sidebar:   #F2DABB;
    --c-panel:     #EDD9A3;
    --c-header:    #4A6BAA;
    --c-header-lt: #6A90D0;
    --c-list-bg:   #FBF9F4;
    --c-accent:    #72CC60;
    --c-accent-dk: #52A840;
    --c-gold:      #D4960A;
    --c-red:       #C84040;
    --c-blue:      #4878B8;
    --c-outline:   #5C3A1E;
    --c-text-dk:   #1A1008;
    --c-text-blue: #1E3468;
    --c-text-lt:   #FFFFFF;
    --c-th-bg:     #3A5898;
    --c-th-top:    #6A90D0;
    --c-row-even:  #FFFFFF;
    --c-row-odd:   #EEF3F8;
    --c-div:       #B8C8DC;
    --c-btn:       #72CC60;
    --c-btn-hov:   #52A840;
    --c-btn-red:   #D04030;
    --c-btn-blue:  #3A64AA;
    --font: 'Nanum Square', 'Malgun Gothic', sans-serif;
}

html, body, [class*="css"] { font-family: var(--font) !important; }

.block-container { padding-top: 0 !important; padding-bottom: 1rem !important; }
header[data-testid="stHeader"] { display: none !important; }
.stDeployButton { display: none !important; }
#MainMenu, footer { visibility: hidden; }

/* ── 최상단 헤더 ── */
.wms-header {
    background: var(--c-header);
    border-bottom: 3px solid var(--c-outline);
    padding: 10px 20px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    box-shadow: 0 3px 8px rgba(0,0,0,0.25);
}
.wms-header-title { color:#fff; font-size:22px; font-weight:800; letter-spacing:2px; }
.wms-header-sub   { color:#A8C8F0; font-size:13px; margin-top:2px; }
.wms-header-right { color:#FFE878; font-size:14px; font-weight:700; }

/* ── 사이드바 ── */
[data-testid="stSidebar"] {
    background: var(--c-sidebar) !important;
    border-right: 3px solid var(--c-outline) !important;
}
[data-testid="stSidebar"] > div { padding-top: 0 !important; }

.sidebar-logo {
    background: var(--c-header);
    padding: 14px 12px 10px;
    margin-bottom: 6px;
    border-bottom: 3px solid var(--c-outline);
    text-align: center;
}
.sidebar-logo-title { color:#fff; font-size:20px; font-weight:800; letter-spacing:3px; }
.sidebar-logo-sub   { color:#A8C8F0; font-size:11px; }
.tag-ok  { background:var(--c-accent); color:#fff; font-size:11px; font-weight:700; border-radius:6px; padding:2px 8px; margin-top:4px; display:inline-block; }
.tag-no  { background:var(--c-red);    color:#fff; font-size:11px; font-weight:700; border-radius:6px; padding:2px 8px; margin-top:4px; display:inline-block; }

/* ── 사이드바 버튼 ── */
div[data-testid="stSidebar"] .stButton > button {
    width: 100% !important;
    background: var(--c-sidebar) !important;
    color: var(--c-text-dk) !important;
    border: none !important;
    border-radius: 0 !important;
    padding: 13px 16px !important;
    text-align: left !important;
    font-size: 15px !important;
    font-weight: 600 !important;
    box-shadow: none !important;
    transition: background 0.12s !important;
}
div[data-testid="stSidebar"] .stButton > button:hover {
    background: #A0CC90 !important;
    color: #fff !important;
}

/* ── 페이지 타이틀 ── */
.page-title {
    font-size: 24px; font-weight: 800;
    color: var(--c-text-blue);
    padding: 12px 16px 6px;
    border-bottom: 2px solid var(--c-outline);
    margin-bottom: 14px;
}

/* ── 카드 ── */
.wms-card {
    background: var(--c-list-bg);
    border: 3px solid var(--c-outline);
    border-radius: 14px;
    padding: 20px 24px;
    margin-bottom: 14px;
    box-shadow: 3px 4px 0 var(--c-outline);
}

/* ── 테이블 ── */
.wms-table { width:100%; border-collapse:collapse; font-size:14px; }
.wms-th {
    background: var(--c-th-bg); color:#fff; font-weight:700;
    padding: 12px 10px; text-align: center;
    border-bottom: 3px solid var(--c-outline);
}
.wms-tr-even { background: var(--c-row-even); }
.wms-tr-odd  { background: var(--c-row-odd); }
.wms-td {
    padding: 10px; text-align: center;
    border-bottom: 1px solid var(--c-div);
    color: var(--c-text-dk); vertical-align: middle;
}
.wms-td-left { text-align:left; padding-left:14px; }
.wms-td-red  { color: var(--c-red); font-weight:700; }
.wms-td-blue { color: var(--c-blue); font-weight:700; }

/* ── 일반 버튼 ── */
.stButton > button {
    background: var(--c-btn) !important;
    color: #fff !important;
    border: 2px solid var(--c-outline) !important;
    border-radius: 22px !important;
    font-weight: 700 !important;
    font-size: 15px !important;
    padding: 10px 24px !important;
    box-shadow: 0 4px 0 var(--c-outline) !important;
    transition: all 0.1s !important;
}
.stButton > button:hover {
    background: var(--c-btn-hov) !important;
    transform: translateY(2px) !important;
    box-shadow: 0 2px 0 var(--c-outline) !important;
}
.stButton > button:active {
    transform: translateY(4px) !important;
    box-shadow: none !important;
}

/* ── 입력 필드 ── */
.stTextInput > div > div > input,
.stNumberInput > div > div > input,
.stTextArea textarea {
    border: 2px solid var(--c-outline) !important;
    border-radius: 10px !important;
    font-family: var(--font) !important;
    background: #fff !important;
    color: var(--c-text-dk) !important;
}
.stSelectbox > div > div {
    border: 2px solid var(--c-outline) !important;
    border-radius: 10px !important;
    background: #fff !important;
}

/* ── 탭 ── */
.stTabs [data-baseweb="tab-list"] {
    background: var(--c-panel);
    border-bottom: 3px solid var(--c-outline);
    gap: 4px;
}
.stTabs [data-baseweb="tab"] {
    background: #A090B8; color:#fff;
    border-radius: 8px 8px 0 0;
    font-weight: 700; font-size: 15px;
    padding: 10px 22px;
}
.stTabs [aria-selected="true"] {
    background: var(--c-list-bg) !important;
    color: var(--c-text-blue) !important;
    border-bottom: 3px solid var(--c-list-bg) !important;
}

/* ── 구분선 ── */
.wms-sep { height:2px; background:var(--c-outline); margin:10px 0; border:none; }

/* ── 로그인 카드 ── */
.login-banner {
    background: var(--c-header);
    border-radius: 16px 16px 0 0;
    padding: 24px 28px 18px;
    border: 3px solid var(--c-outline);
    border-bottom: none;
}
.login-body {
    background: var(--c-list-bg);
    border: 3px solid var(--c-outline);
    border-top: none;
    border-radius: 0 0 16px 16px;
    padding: 28px 32px 24px;
    box-shadow: 4px 6px 0 var(--c-outline);
}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# 최상단 헤더
# ══════════════════════════════════════════════════════════════════
now_str = datetime.now().strftime("📅 %Y-%m-%d  %H:%M")
st.markdown(f"""
<div class="wms-header">
  <div>
    <div class="wms-header-title">📦 WMS SYSTEM</div>
    <div class="wms-header-sub">벤더플렉스 통합 관리 시스템</div>
  </div>
  <div class="wms-header-right">{now_str}</div>
</div>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# 사이드바
# ══════════════════════════════════════════════════════════════════
with st.sidebar:
    sa_ok    = st.session_state.svc_sa    is not None
    oauth_ok = st.session_state.svc_oauth is not None
    both_ok  = sa_ok and oauth_ok

    if both_ok:
        tag = '<span class="tag-ok">✅ 연결됨</span>'
    elif sa_ok or oauth_ok:
        tag = '<span class="tag-no" style="background:#C08010;">⚠️ 일부 연결</span>'
    else:
        tag = '<span class="tag-no">🔒 미연결</span>'

    st.markdown(f"""
    <div class="sidebar-logo">
      <div class="sidebar-logo-title">WMS</div>
      <div class="sidebar-logo-sub">벤더플렉스 v2.0</div>
      {tag}
    </div>
    """, unsafe_allow_html=True)

    if not both_ok:
        sa_txt    = "✅" if sa_ok    else "❌"
        oauth_txt = "✅" if oauth_ok else "❌"
        st.markdown(f"""
        <div style="font-size:12px;padding:8px 12px;color:var(--c-text-dk);">
          {sa_txt} 서비스 계정 (실재고)<br>
          {oauth_txt} 구글 계정 (벤더/창고)
        </div>
        """, unsafe_allow_html=True)

        # SA는 됐지만 OAuth만 없는 경우 → 사이드바에서 바로 로그인 버튼 제공
        if sa_ok and not oauth_ok:
            has_oauth_secrets = ("OAUTH_CLIENT_ID" in st.secrets and "OAUTH_CLIENT_SECRET" in st.secrets)
            if has_oauth_secrets:
                redirect_uri = _get_redirect_uri()
                auth_url = get_oauth_auth_url(redirect_uri)
                st.markdown(f"""
                <div style="padding:8px 12px;">
                  <a href="{auth_url}" target="_self"
                     style="display:block;text-align:center;background:var(--c-btn);
                            color:#fff;font-weight:800;padding:10px 0;border-radius:16px;
                            text-decoration:none;border:2px solid var(--c-outline);
                            box-shadow:0 3px 0 var(--c-outline);font-size:14px;">
                    🔑  구글 계정 로그인
                  </a>
                </div>
                """, unsafe_allow_html=True)
            else:
                if st.button("🔑  인증 설정", key="nav_login"):
                    st.session_state.page = "login"
                    st.rerun()
        else:
            if st.button("🔑  인증 설정", key="nav_login"):
                st.session_state.page = "login"
                st.rerun()
    else:
        pages = [
            ("🏠", "실재고 현황",     "actual"),
            ("📦", "벤더재고 현황",   "vendor"),
            ("🔄", "재고 이동",       "transfer"),
            ("📥", "벤더플렉스 입고", "receive"),
            ("✏️", "출고량 기록",     "dispatch"),
            ("⚖️", "재고 차감/조정",  "adjust"),
            ("📋", "발주서 취합",     "po"),
        ]
        for icon, label, key in pages:
            if st.button(f"{icon}  {label}", key=f"nav_{key}"):
                st.session_state.page = key
                st.rerun()

        st.markdown("<hr style='border-color:#C0A870;margin:8px 0;'>", unsafe_allow_html=True)
        if st.button("🏠  메인화면", key="nav_main"):
            st.session_state.page = "main"
            st.rerun()
        if st.button("🔓  로그아웃", key="nav_logout"):
            st.session_state.svc_sa    = None
            st.session_state.svc_oauth = None
            st.session_state.oauth_token = None
            st.session_state.page = "login"
            st.cache_data.clear()
            st.rerun()


# ── 미인증 시 로그인 강제 이동 ──
# SA도 없으면 무조건 로그인 페이지
# SA는 있고 OAuth만 없는 경우 → 사이드바에서 바로 로그인 버튼 제공, 강제 이동 안 함
_sa_ready    = st.session_state.svc_sa    is not None
_oauth_ready = st.session_state.svc_oauth is not None
if not _sa_ready and st.session_state.page != "login":
    st.session_state.page = "login"
elif _sa_ready and not _oauth_ready and st.session_state.page not in ("login", "main"):
    st.session_state.page = "main"   # 메뉴 접근 차단, 메인은 허용

svc_sa    = st.session_state.svc_sa     # ACTUAL_ID 전용
svc_oauth = st.session_state.svc_oauth  # VENDOR_ID, WAREHOUSE_ID 전용

# ══════════════════════════════════════════════════════════════════
# 라우터
# ══════════════════════════════════════════════════════════════════
page = st.session_state.get("page", "login")

if not _sa_ready:
    page_login()
elif not _oauth_ready:
    # SA는 됐지만 OAuth 미연결 → 메인화면은 보여주되 메뉴 기능 차단
    # 사이드바의 로그인 버튼으로 OAuth 처리
    page_main()
elif page == "login":
    page_main()
elif page == "main":
    page_main()
elif page == "actual":
    page_actual()
elif page == "vendor":
    page_vendor()
elif page == "transfer":
    page_transfer()
elif page == "receive":
    page_receive()
elif page == "dispatch":
    page_dispatch()
elif page == "adjust":
    page_adjust()
elif page == "po":
    page_po()


# ══════════════════════════════════════════════════════════════════
# PAGE: LOGIN
# ══════════════════════════════════════════════════════════════════
def page_login():
    col_l, col_c, col_r = st.columns([1, 2, 1])
    with col_c:
        st.markdown("""
        <div class="login-banner">
          <div style="display:flex;align-items:center;gap:14px;">
            <span style="font-size:38px;">📦</span>
            <div>
              <div style="font-size:24px;font-weight:800;color:#fff;">벤더플렉스 WMS</div>
              <div style="font-size:13px;color:#A8C8F0;">통합 재고 관리 시스템</div>
            </div>
          </div>
        </div>
        <div class="login-body">
        """, unsafe_allow_html=True)

        sa_ok    = st.session_state.svc_sa    is not None
        oauth_ok = st.session_state.svc_oauth is not None

        # ── STEP 1: 서비스 계정 ──────────────────────────────────
        step1_icon = "✅" if sa_ok else "1️⃣"
        sa_method = "Secrets 자동 연결" if "gcp_service_account" in st.secrets else "JSON 파일 업로드"
        st.markdown(f"""
        <div style="font-size:16px;font-weight:800;color:var(--c-text-blue);margin-bottom:4px;">
          {step1_icon}  실재고 시트 — 서비스 계정 연결
        </div>
        <div style="font-size:12px;color:#999;margin-bottom:8px;">
          방식: {sa_method}
        </div>
        """, unsafe_allow_html=True)

        if not sa_ok:
            has_sa_secret = "gcp_service_account" in st.secrets
            if has_sa_secret:
                # Secrets에 있는데 아직 안 됐다면 오류 상황
                st.error("Secrets의 [gcp_service_account] 연결에 실패했습니다. 값을 확인해주세요.")
                if st.button("🔄 재시도", key="retry_sa"):
                    _try_auto_sa()
                    st.rerun()
            else:
                # Secrets 없음 → 파일 업로드 fallback
                st.caption("또는 Secrets에 [gcp_service_account]를 추가하면 자동 연결됩니다.")
                sa_file = st.file_uploader("서비스 계정 키 JSON 업로드", type=["json"], key="sa_json")
                if sa_file:
                    try:
                        creds_dict = json.loads(sa_file.read().decode("utf-8"))
                        with st.spinner("서비스 계정 연결 중..."):
                            service = build_sa_service(creds_dict)
                            service.spreadsheets().get(spreadsheetId=ACTUAL_ID).execute()
                        st.session_state.svc_sa = service
                        st.success("✅ 서비스 계정 연결 성공!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"서비스 계정 연결 실패: {e}")
        else:
            src = "Secrets 자동 연결" if "gcp_service_account" in st.secrets else "JSON 업로드"
            st.success(f"✅ 서비스 계정 연결됨 ({src})")

        st.markdown('<hr class="wms-sep">', unsafe_allow_html=True)

        # ── STEP 2: OAuth (Secrets 기반, 자동 유지) ──────────────
        step2_icon = "✅" if oauth_ok else "2️⃣"
        st.markdown(f"""
        <div style="font-size:16px;font-weight:800;color:var(--c-text-blue);margin-bottom:4px;">
          {step2_icon}  벤더/창고 시트 — 구글 계정 연결
        </div>
        <div style="font-size:12px;color:#999;margin-bottom:8px;">
          Streamlit Secrets의 refresh_token으로 자동 유지 · 최초 1회만 로그인 필요
        </div>
        """, unsafe_allow_html=True)

        if not oauth_ok:
            has_secrets = ("OAUTH_CLIENT_ID" in st.secrets and "OAUTH_CLIENT_SECRET" in st.secrets)
            if not has_secrets:
                st.warning("""
**Streamlit Secrets 설정 필요**

Streamlit Cloud → 앱 설정 → Secrets에 아래를 추가하세요:
```toml
OAUTH_CLIENT_ID     = "..."
OAUTH_CLIENT_SECRET = "..."
```
                """)
            else:
                redirect_uri = _get_redirect_uri()
                auth_url = get_oauth_auth_url(redirect_uri)
                st.markdown(f"""
                <div style="background:#E8F4FF;border:2px solid var(--c-blue);
                            border-radius:10px;padding:14px;margin-top:4px;">
                  <div style="font-weight:800;color:var(--c-blue);margin-bottom:8px;">
                    🔗 구글 계정으로 로그인 (최초 1회)
                  </div>
                  <a href="{auth_url}" target="_self"
                     style="display:inline-block;background:var(--c-btn);color:#fff;
                            font-weight:800;padding:10px 24px;border-radius:20px;
                            text-decoration:none;border:2px solid var(--c-outline);
                            box-shadow:0 3px 0 var(--c-outline);">
                    🔑  구글 계정으로 로그인
                  </a>
                  <div style="font-size:11px;color:#888;margin-top:8px;">
                    로그인 후 발급된 refresh_token을 Secrets에 저장하면 이후 자동 연결됩니다.
                  </div>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.success("✅ 구글 계정 연결됨 (자동)")

        # ── 최초 로그인 후 토큰 저장 안내 ──────────────────────────
        new_token = st.session_state.get("_new_token_dict")
        if new_token and not st.session_state.get("_first_token_shown"):
            st.session_state["_first_token_shown"] = True
            refresh = new_token.get("refresh_token", "")
            st.markdown(f"""
            <div style="background:#FFF8E0;border:2px solid var(--c-gold);
                        border-radius:10px;padding:14px;margin-top:10px;">
              <div style="font-weight:800;color:var(--c-gold);margin-bottom:6px;">
                🔑 최초 로그인 완료! Secrets에 아래 값을 추가하세요
              </div>
              <div style="font-size:12px;color:#555;margin-bottom:8px;">
                Streamlit Cloud → 앱 설정 → Secrets → 아래 줄 추가 후 앱 재시작<br>
                이후부터는 구글 계정 로그인 없이 자동 연결됩니다.
              </div>
              <code style="display:block;background:#f5f0e0;padding:10px;
                           border-radius:6px;font-size:12px;word-break:break-all;">
OAUTH_REFRESH_TOKEN = "{refresh}"
              </code>
            </div>
            """, unsafe_allow_html=True)

        st.markdown('<hr class="wms-sep">', unsafe_allow_html=True)

        if sa_ok and oauth_ok:
            st.success("🎀 모든 인증 완료!")
            if st.button("🚀  WMS 시작하기", key="go_main"):
                st.session_state.page = "main"
                st.rerun()
        else:
            missing = []
            if not sa_ok:    missing.append("① 서비스 계정 JSON 업로드")
            if not oauth_ok: missing.append("② 구글 계정 로그인")
            st.info(f"남은 단계: {' / '.join(missing)}")

        st.markdown("""
          <div style="margin-top:16px;padding-top:12px;border-top:2px solid var(--c-div);
                      font-size:11px;color:#A09080;text-align:center;">
            벤더플렉스 WMS  v2.0  ·  Streamlit Edition
          </div>
        </div>
        """, unsafe_allow_html=True)


def page_main():
    st.markdown('<div class="page-title">📦 WMS 메인 화면</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="wms-card">
      <div style="display:flex;align-items:center;gap:18px;margin-bottom:14px;">
        <span style="font-size:46px;">📦</span>
        <div>
          <div style="font-size:24px;font-weight:800;color:var(--c-text-blue);">벤더플렉스 WMS</div>
          <div style="font-size:14px;color:#4A6890;">통합 재고 관리 시스템</div>
        </div>
      </div>
      <hr class="wms-sep">
      <div style="display:flex;flex-direction:column;gap:10px;padding-top:6px;">
        <div style="display:flex;align-items:center;gap:12px;"><span style="font-size:20px;">🏠</span><span style="font-weight:700;">벤더플렉스 WMS에 오신 것을 환영합니다.</span></div>
        <div style="display:flex;align-items:center;gap:12px;"><span style="font-size:20px;">📋</span><span style="font-weight:700;">왼쪽 메뉴에서 원하는 기능을 선택하세요.</span></div>
        <div style="display:flex;align-items:center;gap:12px;"><span style="font-size:20px;">🔗</span><span style="font-weight:700;">재고 현황은 실시간으로 Google Sheets와 연동됩니다.</span></div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("### 빠른 이동")
    cols = st.columns(4)
    quick = [
        ("📊", "실재고 현황",     "actual"),
        ("📦", "벤더재고 현황",   "vendor"),
        ("📥", "벤더플렉스 입고", "receive"),
        ("✏️", "출고량 기록",     "dispatch"),
        ("⚖️", "재고 차감/조정",  "adjust"),
        ("📋", "발주서 취합",     "po"),
    ]
    for i, (icon, label, key) in enumerate(quick):
        with cols[i % 4]:
            if st.button(f"{icon}  {label}", key=f"quick_{key}"):
                st.session_state.page = key
                st.rerun()


# ══════════════════════════════════════════════════════════════════
# PAGE: ACTUAL INVENTORY (실재고 현황)
# ══════════════════════════════════════════════════════════════════
def page_actual():
    st.markdown('<div class="page-title">📊 실재고 현황</div>', unsafe_allow_html=True)
    col_upd, col_btn = st.columns([6, 1])
    with col_btn:
        if st.button("🔄 새로고침", key="actual_refresh"):
            st.cache_data.clear()

    @st.cache_data(ttl=120, show_spinner="📊 데이터 불러오는 중...")
    def load_actual(_key):
        sm = svc_oauth.spreadsheets().get(spreadsheetId=ACTUAL_ID).execute()
        sn = sm["sheets"][0]["properties"]["title"]
        tr = svc_oauth.spreadsheets().values().get(spreadsheetId=ACTUAL_ID, range=f"'{sn}'!J1").execute()
        j1 = tr.get("values", [])
        upd_txt = j1[0][0] if j1 and j1[0] else "업데이트 기록 없음"

        vr = svc_oauth.spreadsheets().values().get(spreadsheetId=VENDOR_ID, range="'벤더플렉스 출고량'!D3:E").execute()
        raw = vr.get("values", [])
        vm = svc_oauth.spreadsheets().get(spreadsheetId=VENDOR_ID, ranges=["'벤더플렉스 출고량'!D3:D"], includeGridData=True).execute()
        rmeta = []
        try: rmeta = vm["sheets"][0]["data"][0].get("rowMetadata", [])
        except Exception: pass
        base2 = [r for i, r in enumerate(raw) if not (i < len(rmeta) and (rmeta[i].get("hiddenByUser") or rmeta[i].get("hiddenByFilter")))]

        today = datetime.now()
        h2 = svc_oauth.spreadsheets().values().get(spreadsheetId=WAREHOUSE_ID, range="창고별실재고!1:2", valueRenderOption="FORMATTED_VALUE").execute()
        hrows = h2.get("values", [])
        row1 = hrows[0] if hrows else []
        row2 = hrows[1] if len(hrows) > 1 else []

        def fdc(d):
            fmts = [f"{d.year}. {d.month}. {d.day}", f"{d.year}.{d.month}.{d.day}", f"{d.year}. {d.month:02d}. {d.day:02d}", f"{d.year}.{d.month:02d}.{d.day:02d}"]
            return next((i for i, c in enumerate(row1) if str(c).strip() in fmts), -1)

        ci = fdc(today)
        if ci == -1: ci = fdc(today - timedelta(days=1))

        wh = {}
        if ci != -1:
            jk = ci + 3; jp = ci + 4
            for i in range(ci, min(ci + 6, len(row2))):
                h = str(row2[i]).replace("\n", "").replace(" ", "")
                if "적재" in h or "2적" in h: jk = i
                if "집품" in h or "2집" in h: jp = i
            jkl = col_letter(jk); jpl = col_letter(jp)
            rd = svc_oauth.spreadsheets().values().batchGet(
                spreadsheetId=WAREHOUSE_ID,
                ranges=["창고별실재고!B3:B", f"창고별실재고!{jpl}3:{jpl}", f"창고별실재고!{jkl}3:{jkl}"],
                valueRenderOption="UNFORMATTED_VALUE",
            ).execute()
            vrs = rd.get("valueRanges", [])
            wi = vrs[0].get("values", [])
            jd = vrs[1].get("values", []) if len(vrs) > 1 else []
            kd = vrs[2].get("values", []) if len(vrs) > 2 else []
            for i, r in enumerate(wi):
                if not r: continue
                nm = str(r[0]).strip()
                wh[nm] = {
                    "jp": str(jd[i][0]) if i < len(jd) and jd[i] else "0",
                    "jk": str(kd[i][0]) if i < len(kd) and kd[i] else "0",
                }
        return upd_txt, base2, wh, ci

    try:
        upd_txt, base2, wh, ci = load_actual(str(int(datetime.now().timestamp() // 120)))
    except Exception as e:
        st.error(f"데이터 로드 실패: {e}"); return

    with col_upd:
        st.caption(f"마지막 업데이트: {upd_txt}")
    if ci == -1:
        st.warning("시트에서 오늘/어제 날짜 열을 찾지 못했습니다.")

    items_data = st.session_state.get("items_data", {})
    rows_html = ""
    for di, row in enumerate(base2 or []):
        if not row: continue
        sku = str(row[0]).strip()
        if not sku or sku == "-": continue
        name = str(row[1]).strip() if len(row) > 1 else sku
        w2 = wh.get(sku, {"jp": "0", "jk": "0"})
        bad = items_data.get(sku, {}).get("bad", "0")
        cls = "wms-tr-even" if di % 2 == 0 else "wms-tr-odd"
        rows_html += f'<tr class="{cls}"><td class="wms-td wms-td-left">{name}</td><td class="wms-td">{w2["jp"]}</td><td class="wms-td">{w2["jk"]}</td><td class="wms-td">{bad}</td></tr>'

    st.markdown(f"""
    <div class="wms-card" style="padding:0;overflow:auto;">
    <table class="wms-table">
      <thead><tr>
        <th class="wms-th" style="text-align:left;padding-left:14px;">품목명</th>
        <th class="wms-th">2집</th><th class="wms-th">2적</th><th class="wms-th">불용</th>
      </tr></thead>
      <tbody>{rows_html}</tbody>
    </table></div>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# PAGE: VENDOR INVENTORY (벤더재고 현황)
# ══════════════════════════════════════════════════════════════════
def page_vendor():
    st.markdown('<div class="page-title">📦 벤더재고 현황</div>', unsafe_allow_html=True)

    c_date, c_prev, c_next, c_ref = st.columns([3, 1, 1, 1])
    with c_date:
        sel = st.date_input("날짜", value=st.session_state.vendor_date.date(), label_visibility="collapsed")
        if sel: st.session_state.vendor_date = datetime(sel.year, sel.month, sel.day)
    with c_prev:
        if st.button("◀", key="v_prev"): st.session_state.vendor_date -= timedelta(days=1); st.rerun()
    with c_next:
        if st.button("▶", key="v_next"): st.session_state.vendor_date += timedelta(days=1); st.rerun()
    with c_ref:
        if st.button("🔄", key="v_ref"): st.cache_data.clear(); st.rerun()

    cur = st.session_state.vendor_date

    @st.cache_data(ttl=120, show_spinner="📦 벤더 재고 불러오는 중...")
    def load_vendor(_ds):
        d = datetime.strptime(_ds, "%Y-%m-%d")
        ts = f"{d.year}. {d.month}. {d.day}"
        hr = svc_oauth.spreadsheets().values().get(spreadsheetId=VENDOR_ID, range="벤더플렉스 출고량!1:1", valueRenderOption="FORMATTED_VALUE").execute()
        hrow = hr.get("values", [[]])[0]
        ci = next((i for i, c in enumerate(hrow) if str(c).strip() == ts), -1)
        new_date = None
        if ci == -1:
            yd = d - timedelta(days=1); ts2 = f"{yd.year}. {yd.month}. {yd.day}"
            ci2 = next((i for i, c in enumerate(hrow) if str(c).strip() == ts2), -1)
            if ci2 != -1: ci = ci2; new_date = yd.strftime("%Y-%m-%d"); ts = ts2
        if ci == -1: return None, None, None, ts
        cs = col_letter(ci); ce = col_letter(ci + 2)
        res = svc_oauth.spreadsheets().values().batchGet(
            spreadsheetId=VENDOR_ID,
            ranges=["벤더플렉스 출고량!E3:E", f"벤더플렉스 출고량!{cs}3:{ce}"],
            valueRenderOption="UNFORMATTED_VALUE",
        ).execute()
        vrs = res.get("valueRanges", [])
        id_ = vrs[0].get("values", [])
        inv = vrs[1].get("values", []) if len(vrs) > 1 else []
        total = 0; rows = []
        for i, ir in enumerate(id_):
            if not ir: continue
            nm = ir[0]; j = ip = ch = 0
            if i < len(inv):
                r = inv[i]
                try: j  = float(r[0]) if r else 0
                except: pass
                try: ip = float(r[1]) if len(r) > 1 else 0
                except: pass
                try: ch = float(r[2]) if len(r) > 2 else 0
                except: pass
            if j == 0 and ch == 0 and ip == 0: continue
            total += ch
            def _fmt(n):
                try: return str(int(n)) if float(n) == int(float(n)) else str(n)
                except: return str(n)
            rows.append((nm, _fmt(ip), _fmt(ch), _fmt(j)))
        return rows, total, new_date, ts

    try:
        rows, total, new_date, ts = load_vendor(cur.strftime("%Y-%m-%d"))
    except Exception as e:
        st.error(f"데이터 로드 실패: {e}"); return

    if new_date:
        st.session_state.vendor_date = datetime.strptime(new_date, "%Y-%m-%d")
        st.info(f"해당 날짜 데이터가 없어 {new_date} 데이터를 표시합니다.")
    if rows is None:
        st.error(f"해당 날짜({ts})의 데이터를 찾을 수 없습니다."); return

    st.markdown(f"<div style='color:var(--c-red);font-size:17px;font-weight:800;margin-bottom:10px;'>총 출고량: {int(total) if total else 0}</div>", unsafe_allow_html=True)
    rows_html = "".join(
        f'<tr class="{"wms-tr-even" if di%2==0 else "wms-tr-odd"}"><td class="wms-td wms-td-left">{nm}</td><td class="wms-td">{ip}</td><td class="wms-td">{ch}</td><td class="wms-td">{jg}</td></tr>'
        for di, (nm, ip, ch, jg) in enumerate(rows)
    )
    st.markdown(f"""
    <div class="wms-card" style="padding:0;overflow:auto;">
    <table class="wms-table">
      <thead><tr>
        <th class="wms-th" style="text-align:left;padding-left:14px;">상품명</th>
        <th class="wms-th">입고</th><th class="wms-th">출고</th><th class="wms-th">재고</th>
      </tr></thead>
      <tbody>{rows_html}</tbody>
    </table></div>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# PAGE: TRANSFER (재고 이동)
# ══════════════════════════════════════════════════════════════════
def page_transfer():
    st.markdown('<div class="page-title">🔄 재고 이동</div>', unsafe_allow_html=True)
    st.info("🛠️ 해당 메뉴는 업데이트 준비 중입니다! 조금만 기다려 주세요. 🎀")


# ══════════════════════════════════════════════════════════════════
# PAGE: RECEIVE INVENTORY (벤더플렉스 입고) — PDF 기능 제외
# ══════════════════════════════════════════════════════════════════
def page_receive():
    st.markdown('<div class="page-title">📥 벤더플렉스 입고</div>', unsafe_allow_html=True)

    today = datetime.now()
    st.markdown(f"<div style='color:var(--c-red);font-weight:800;font-size:15px;margin-bottom:12px;'>📅 적용 일자: {today.strftime('%Y. %m. %d')}</div>", unsafe_allow_html=True)

    @st.cache_data(ttl=300, show_spinner=False)
    def load_lim():
        res = svc_oauth.spreadsheets().values().get(spreadsheetId=VENDOR_ID, range="벤더플렉스 출고량!B3:E").execute()
        return [{"loc": str(r[0]).strip(), "item": str(r[3]).strip()} for r in res.get("values", []) if len(r) > 3]

    try:
        lim = load_lim()
    except Exception as e:
        st.error(f"로케이션 데이터 로드 실패: {e}"); return

    num_rows = st.number_input("입고 행 수", min_value=5, max_value=50, value=15, step=5)

    h1, h2, h3 = st.columns([2, 5, 2])
    h1.markdown(f"<div style='font-weight:800;color:var(--c-text-blue);font-size:14px;'>{LOC_PREFIX}**번호**</div>", unsafe_allow_html=True)
    h2.markdown("<div style='font-weight:800;color:var(--c-text-blue);font-size:14px;'>품목명</div>", unsafe_allow_html=True)
    h3.markdown("<div style='font-weight:800;color:var(--c-text-blue);font-size:14px;'>입고수량</div>", unsafe_allow_html=True)
    st.markdown('<hr class="wms-sep" style="margin:4px 0 8px;">', unsafe_allow_html=True)

    recv_rows = []
    for i in range(int(num_rows)):
        c1, c2, c3 = st.columns([2, 5, 2])
        with c1:
            loc_val = st.text_input("loc", key=f"recv_loc_{i}", placeholder="예: 01", label_visibility="collapsed")
        with c2:
            item_default = ""
            if loc_val:
                full = f"{LOC_PREFIX}{loc_val}"
                matches = [m for m in lim if full == m["loc"]] or [m for m in lim if loc_val in m["loc"]]
                item_default = matches[0]["item"] if matches else "확인 요망 (미등록)"
            item_val = st.text_input("item", value=item_default, key=f"recv_item_{i}", placeholder="품목명", label_visibility="collapsed")
        with c3:
            qty_val = st.text_input("qty", key=f"recv_qty_{i}", placeholder="수량", label_visibility="collapsed")
        if loc_val or item_val:
            recv_rows.append({"item": item_val, "qty": qty_val})

    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("🚀  작성된 내용 일괄 입고하기", key="recv_submit"):
        upd = {}
        for r in recv_rows:
            item = r["item"].strip()
            if "로케이션" in item: st.error(f"잘못된 품목명: {item}"); return
            q = _safe_int(r["qty"].strip(), default=None)
            if item and q is not None and q > 0: upd[item] = upd.get(item, 0) + q
        if not upd: st.warning("입고할 수량이 없습니다!"); return

        ts = f"{today.year}. {today.month}. {today.day}"
        try:
            with st.spinner("시트에 입고 기록 중... 🚀"):
                hr = svc_oauth.spreadsheets().values().get(spreadsheetId=VENDOR_ID, range="벤더플렉스 출고량!1:1", valueRenderOption="FORMATTED_VALUE").execute()
                hrow = hr.get("values", [[]])[0]
                ci = next((i for i, c in enumerate(hrow) if str(c).strip() == ts), -1)
                if ci == -1: st.error(f"시트에서 해당 날짜({ts})를 찾을 수 없어요!"); return
                il = col_letter(ci + 1)
                res = svc_oauth.spreadsheets().values().batchGet(
                    spreadsheetId=VENDOR_ID,
                    ranges=["벤더플렉스 출고량!E:E", f"벤더플렉스 출고량!{il}:{il}"],
                    valueRenderOption="UNFORMATTED_VALUE",
                ).execute()
                vrs = res.get("valueRanges", [])
                ed = vrs[0].get("values", [])
                id2 = vrs[1].get("values", []) if len(vrs) > 1 else []
                batch = []; msg = ""
                for item, add in upd.items():
                    ri = next((i for i, r in enumerate(ed) if r and str(r[0]).strip() == item), -1)
                    if ri != -1:
                        cur = _safe_int(id2[ri][0]) if ri < len(id2) and id2[ri] else 0
                        nw = cur + add
                        batch.append({"range": f"'벤더플렉스 출고량'!{il}{ri+1}", "values": [[nw]]})
                        msg += f"[{item}] {cur} + {add} = {nw}개\n"
                if batch:
                    svc_oauth.spreadsheets().values().batchUpdate(spreadsheetId=VENDOR_ID, body={"valueInputOption": "USER_ENTERED", "data": batch}).execute()
                    st.success(f"🎀 입고 완료!\n\n{msg}")
                    load_lim.clear()
        except Exception as e:
            st.error(f"구글 시트 기록 중 문제: {e}")


# ══════════════════════════════════════════════════════════════════
# PAGE: DISPATCH (출고량 기록)
# ══════════════════════════════════════════════════════════════════
def page_dispatch():
    st.markdown('<div class="page-title">✏️ 출고량 기록</div>', unsafe_allow_html=True)
    st.caption("💡 마감 재고를 입력하면 출고량이 자동 계산됩니다.")

    tgt = (datetime.now() - timedelta(hours=6)).date()

    @st.cache_data(ttl=60, show_spinner="✏️ 출고량 데이터 분석 중...")
    def load_dispatch(_ts):
        poss = [f"{tgt.year}.{tgt.month}.{tgt.day}", f"{tgt.year}.{tgt.month:02d}.{tgt.day:02d}", tgt.strftime("%m월%d일"), f"{tgt.month}월{tgt.day}일", f"{tgt.year}/{tgt.month}/{tgt.day}", f"{tgt.year}-{tgt.month:02d}-{tgt.day:02d}", f"{tgt.month}.{tgt.day}", f"{tgt.month:02d}.{tgt.day:02d}", str(tgt)]
        res = svc_oauth.spreadsheets().values().get(spreadsheetId=VENDOR_ID, range="'벤더플렉스 출고량'", valueRenderOption="FORMATTED_VALUE").execute()
        sd = res.get("values", [])
        if not sd: raise ValueError("시트 데이터가 비어있어요!")
        tidx = -1; hrow_idx = 0
        for hr_i in range(min(3, len(sd))):
            row1 = sd[hr_i]
            tidx = next((i for i, c in enumerate(row1) if str(c).replace(" ", "").strip() in poss), -1)
            if tidx != -1: hrow_idx = hr_i; break
        row1 = sd[hrow_idx]
        if tidx == -1: raise ValueError(f"날짜를 찾을 수 없어요!\n찾는 날짜: {poss[0]}")
        pi2 = []
        for i in range(tidx - 1, 4, -1):
            if len(pi2) >= 7: break
            if str(row1[i]).strip(): pi2.append(i)
        ditems = []; avg_map = {}
        for row in sd[2:]:
            if len(row) < 5: continue
            loc = str(row[1]).strip(); item = str(row[4]).strip()
            if not item or item == "-": continue
            pre = _safe_int(row[tidx]) if len(row) > tidx else 0
            vd = sum(1 for p in pi2 if len(row) > p and str(row[p]).replace(",", "").lstrip("-").isdigit())
            tp2 = sum(max(0, _safe_int(row[p])) for p in pi2 if len(row) > p and str(row[p]).replace(",", "").lstrip("-").isdigit())
            avg = round(tp2 / vd) if vd > 0 else 0
            avg_map[item] = avg
            ditems.append({"loc": loc, "item": item, "pre": pre})
        return tidx, avg_map, ditems

    try:
        tidx, avg_map, ditems = load_dispatch(str(tgt))
    except Exception as e:
        st.error(f"데이터 분석 실패: {e}"); return

    st.markdown(f"**{tgt.strftime('%Y년 %m월 %d일')} 기준 · 총 {len(ditems)}개 품목**")

    hc = st.columns([1.5, 4, 2, 2, 1.5])
    for col_el, lbl in zip(hc, ["**로케이션**", "**품목명**", "**마감 재고 입력**", "**오늘 출고량**", "**전일 재고**"]):
        col_el.markdown(lbl)
    st.markdown('<hr class="wms-sep" style="margin:4px 0 8px;">', unsafe_allow_html=True)

    dispatch_inputs = {}
    for i, rd in enumerate(ditems):
        c1, c2, c3, c4, c5 = st.columns([1.5, 4, 2, 2, 1.5])
        c1.markdown(f"<div style='padding:10px 4px;font-size:13px;color:#555;'>{rd['loc']}</div>", unsafe_allow_html=True)
        c2.markdown(f"<div style='padding:10px 4px;font-size:14px;font-weight:700;'>{rd['item']}</div>", unsafe_allow_html=True)
        with c3:
            val = st.text_input("", key=f"disp_{i}", placeholder="수량", label_visibility="collapsed")
            dispatch_inputs[i] = val
        if val and val.strip().isdigit():
            out = rd["pre"] - int(val.strip())
            color = "var(--c-red)" if out >= 0 else "#CC0000"
            c4.markdown(f"<div style='padding:10px 4px;font-size:15px;font-weight:800;color:{color};'>{out}</div>", unsafe_allow_html=True)
        else:
            c4.markdown("<div style='padding:10px 4px;color:#aaa;'>-</div>", unsafe_allow_html=True)
        c5.markdown(f"<div style='padding:10px 4px;font-size:13px;color:#888;'>{rd['pre']}</div>", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("🚀  일괄 마감 처리하기", key="dispatch_submit"):
        batch = []; warns = []
        for i, rd in enumerate(ditems):
            v = dispatch_inputs.get(i, "").strip()
            if not v: st.error(f"[{rd['item']}] 마감 재고 칸이 비어있습니다!"); return
            if not v.isdigit(): st.error(f"[{rd['item']}] 수량은 자연수만 입력해 주세요!"); return
            tot = int(v); out = rd["pre"] - tot
            if out < 0: st.error(f"[{rd['item']}] 출고량이 마이너스({out}개)가 될 수 없습니다!"); return
            if avg_map.get(rd["item"], 0) > 0 and tot < avg_map[rd["item"]] * 2: warns.append(rd["item"])
            batch.append({"range": f"'벤더플렉스 출고량'!{col_letter(tidx)}{i+3}", "values": [[tot]]})
        if not batch: st.info("입력된 수량이 없습니다!"); return
        try:
            with st.spinner("시트에 출고 마감 기록 중..."):
                svc_oauth.spreadsheets().values().batchUpdate(spreadsheetId=VENDOR_ID, body={"valueInputOption": "USER_ENTERED", "data": batch}).execute()
            if warns: st.warning(f"마감 완료! 결품 위험:\n👉 {', '.join(warns)}")
            else: st.success("🌸 출고량 기록이 완벽하게 마감되었습니다!")
            load_dispatch.clear()
        except Exception as e:
            st.error(f"기록 중 문제: {e}")


# ══════════════════════════════════════════════════════════════════
# PAGE: ADJUST INVENTORY (재고 차감/조정)
# ══════════════════════════════════════════════════════════════════
def page_adjust():
    st.markdown('<div class="page-title">⚖️ 재고 차감/조정</div>', unsafe_allow_html=True)

    @st.cache_data(ttl=300, show_spinner=False)
    def load_lim_adj():
        res = svc_oauth.spreadsheets().values().get(spreadsheetId=VENDOR_ID, range="벤더플렉스 출고량!B3:E").execute()
        return [{"loc": str(r[0]).strip(), "item": str(r[3]).strip()} for r in res.get("values", []) if len(r) > 3]

    try:
        lim = load_lim_adj()
    except Exception as e:
        st.error(f"데이터 로드 실패: {e}"); return

    def do_adjust(mode, item, qty_str, reason):
        if not item or "요망" in item: st.error("품목명을 확인해 주세요!"); return False
        if not qty_str.isdigit() or int(qty_str) <= 0: st.error("수량은 자연수로 적어주세요!"); return False
        qty = int(qty_str); mul = -1 if mode == "deduct" else 1; add = qty * mul
        today = datetime.now()
        poss = [f"{today.year}.{today.month}.{today.day}", f"{today.year}.{today.month:02d}.{today.day:02d}", today.strftime("%m월%d일"), f"{today.month}월{today.day}일"]
        try:
            meta = svc_oauth.spreadsheets().get(spreadsheetId=WAREHOUSE_ID).execute()
            sid = next((s["properties"]["sheetId"] for s in meta.get("sheets", []) if s["properties"]["title"] == "집품창고입출고및조정"), None)
            hr = svc_oauth.spreadsheets().values().get(spreadsheetId=WAREHOUSE_ID, range="'집품창고입출고및조정'!1:2", valueRenderOption="FORMATTED_VALUE").execute()
            r1 = hr.get("values", [[]])[0]
            dci = next((i for i, c in enumerate(r1) if str(c).replace(" ", "").strip() in poss), -1)
            if dci == -1: st.error("'집품창고입출고및조정'에서 오늘 날짜를 찾을 수 없어요!"); return False
            tci = dci + 5; tcl = col_letter(tci)
            bd = svc_oauth.spreadsheets().values().get(spreadsheetId=WAREHOUSE_ID, range="'집품창고입출고및조정'!B:B").execute()
            bd_ = bd.get("values", [])
            ri = next((i + 1 for i, r in enumerate(bd_) if r and str(r[0]).strip() == item), -1)
            if ri == -1: st.error(f"B열에서 '{item}'을(를) 찾을 수 없습니다!"); return False
            cr = f"'집품창고입출고및조정'!{tcl}{ri}"
            vr = svc_oauth.spreadsheets().values().get(spreadsheetId=WAREHOUSE_ID, range=cr).execute()
            vd = vr.get("values", [[0]]); cur = _safe_int(vd[0][0]) if vd and vd[0] else 0
            nw = cur + add
            svc_oauth.spreadsheets().values().update(spreadsheetId=WAREHOUSE_ID, range=cr, valueInputOption="USER_ENTERED", body={"values": [[nw]]}).execute()
            if sid is not None:
                note = f"[{today.strftime('%Y-%m-%d')} 기록]\n사유: {reason}\n변동: {add}개"
                svc_oauth.spreadsheets().batchUpdate(spreadsheetId=WAREHOUSE_ID, body={"requests": [{"updateCells": {"range": {"sheetId": sid, "startRowIndex": ri-1, "endRowIndex": ri, "startColumnIndex": tci, "endColumnIndex": tci+1}, "rows": [{"values": [{"note": note}]}], "fields": "note"}}]}).execute()

            st.session_state.adjust_log.insert(0, {"time": today.strftime("%Y-%m-%d %H:%M:%S"), "item": item, "reason": reason, "qty": f"+{add}" if add > 0 else str(add)})

            if mode == "deduct" and reason == "파손":
                if item not in st.session_state.items_data: st.session_state.items_data[item] = {"bad": "0"}
                st.session_state.items_data[item]["bad"] = str(int(st.session_state.items_data[item].get("bad", "0")) + qty)
                st.success(f"🎀 [{item}] 차감 완료!\n👉 '불용' 창고에 {qty}개 자동 추가됨")
            else:
                an = "차감" if mode == "deduct" else "조정(추가)"
                st.success(f"🎀 [{item}] 성공적으로 {an} 처리되었습니다! (사유: {reason})")
            return True
        except Exception as e:
            st.error(f"기록 중 문제: {e}"); return False

    tab_deduct, tab_add, tab_log = st.tabs(["  📉  재고 차감  (-)  ", "  📈  재고 조정  (+)  ", "  📝  조정 로그  "])

    with tab_deduct:
        st.markdown('<div class="wms-card">', unsafe_allow_html=True)
        d_loc  = st.text_input("📍 로케이션 번호", key="d_loc", placeholder="번호 입력 (예: 01)")
        d_full = f"{LOC_PREFIX}{d_loc}" if d_loc else ""
        d_m    = ([m for m in lim if d_full == m["loc"]] or [m for m in lim if d_loc in m["loc"]]) if d_loc else []
        d_item = st.text_input("🏷️ 품목명", key="d_item", value=d_m[0]["item"] if d_m else "", placeholder="로케이션 입력 시 자동완성")
        d_qty  = st.text_input("📦 차감 수량", key="d_qty", placeholder="숫자 입력")
        d_rsel = st.selectbox("💬 차감 사유", ["파손", "오출고", "샘플 이동", "자가사용", "폐기", "직접 입력"], key="d_rsel")
        d_rcustom = st.text_input("사유 직접 입력", key="d_rcustom") if d_rsel == "직접 입력" else ""
        if st.button("🚀  차감 실행하기", key="d_submit"):
            with st.spinner("처리 중..."): do_adjust("deduct", d_item.strip(), d_qty.strip(), d_rcustom if d_rsel == "직접 입력" else d_rsel)
        st.markdown('</div>', unsafe_allow_html=True)

    with tab_add:
        st.markdown('<div class="wms-card">', unsafe_allow_html=True)
        a_loc  = st.text_input("📍 로케이션 번호", key="a_loc", placeholder="번호 입력")
        a_full = f"{LOC_PREFIX}{a_loc}" if a_loc else ""
        a_m    = ([m for m in lim if a_full == m["loc"]] or [m for m in lim if a_loc in m["loc"]]) if a_loc else []
        a_item = st.text_input("🏷️ 품목명", key="a_item", value=a_m[0]["item"] if a_m else "", placeholder="로케이션 입력 시 자동완성")
        a_qty  = st.text_input("📦 추가 수량", key="a_qty", placeholder="숫자 입력")
        a_rsel = st.selectbox("💬 추가 사유", ["오출고", "재고 복구", "직접 입력"], key="a_rsel")
        a_rcustom = st.text_input("사유 직접 입력", key="a_rcustom") if a_rsel == "직접 입력" else ""
        if st.button("🚀  추가 실행하기", key="a_submit"):
            with st.spinner("처리 중..."): do_adjust("add", a_item.strip(), a_qty.strip(), a_rcustom if a_rsel == "직접 입력" else a_rsel)
        st.markdown('</div>', unsafe_allow_html=True)

    with tab_log:
        logs = st.session_state.get("adjust_log", [])
        if not logs:
            st.info("📭 기록된 조정 로그가 없습니다.")
        else:
            rows_html = "".join(
                f'<tr class="{"wms-tr-even" if di%2==0 else "wms-tr-odd"}">'
                f'<td class="wms-td">{lg.get("time","")}</td>'
                f'<td class="wms-td wms-td-left">{lg.get("item","")}</td>'
                f'<td class="wms-td">{lg.get("reason","")}</td>'
                f'<td class="wms-td {"wms-td-red" if "-" in str(lg.get("qty","")) else "wms-td-blue"}">{lg.get("qty","")}</td>'
                f'</tr>'
                for di, lg in enumerate(logs)
            )
            st.markdown(f"""
            <div class="wms-card" style="padding:0;overflow:auto;">
            <table class="wms-table">
              <thead><tr>
                <th class="wms-th">시간</th><th class="wms-th">품목명</th>
                <th class="wms-th">사유</th><th class="wms-th">수량</th>
              </tr></thead>
              <tbody>{rows_html}</tbody>
            </table></div>
            """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# PAGE: PO CONSOLIDATION (발주서 취합)
# ══════════════════════════════════════════════════════════════════
def page_po():
    st.markdown('<div class="page-title">📋 발주서 취합 현황</div>', unsafe_allow_html=True)

    # ── 바코드 생성 ──
    st.markdown("### 🔖 하차 바코드 생성")
    ibn_val = st.text_input("IBN 번호 입력 (IBN 뒤 숫자)", key="ibn_input", placeholder="숫자만 입력")
    full_code = f"IBN{ibn_val}" if ibn_val else ""

    if full_code:
        try:
            import barcode
            from barcode.writer import ImageWriter
            bc = barcode.get("code128", full_code, writer=ImageWriter())
            fp = io.BytesIO()
            bc.write(fp, {"module_width": 0.3, "module_height": 8.0, "write_text": False, "quiet_zone": 2.0})
            fp.seek(0)
            st.image(fp, caption=full_code, width=420)
        except Exception as e:
            st.error(f"바코드 생성 실패: {e}")
    else:
        st.markdown('<div style="text-align:center;padding:28px;color:#aaa;font-size:17px;border:2px dashed #ccc;border-radius:10px;margin-bottom:10px;">[ 바코드 대기 중 ]</div>', unsafe_allow_html=True)

    st.markdown("---")

    # ── ZIP 취합 ──
    st.markdown("### 📁 발주서 ZIP 파일 취합")
    uploaded_zip = st.file_uploader("ZIP 파일 선택", type=["zip"], key="po_zip")

    if uploaded_zip:
        try:
            with st.spinner("🧐 데이터 분석 중..."):
                am = svc_oauth.spreadsheets().get(spreadsheetId=ACTUAL_ID).execute()
                an = am["sheets"][0]["properties"]["title"]
                amap = {}
                for r in svc_oauth.spreadsheets().values().get(spreadsheetId=ACTUAL_ID, range=f"'{an}'!A2:G").execute().get("values", []):
                    if r: amap[str(r[0]).strip()] = [str(r[i]).strip() if len(r) > i else "0" for i in range(3, 7)]

                vmap = {}
                for r in svc_oauth.spreadsheets().values().get(spreadsheetId=VENDOR_ID, range="'벤더플렉스 출고량'!B3:D").execute().get("values", []):
                    loc = str(r[0]).strip() if r else ""; sku = str(r[2]).strip() if len(r) > 2 else ""
                    if sku and loc: vmap[sku] = loc

                import openpyxl
                ext = []
                with zipfile.ZipFile(io.BytesIO(uploaded_zip.read()), "r") as z:
                    for fi in z.infolist():
                        if fi.filename.endswith(".xlsx") and not fi.filename.startswith("~"):
                            fn = os.path.basename(fi.filename)
                            ab = "-"; m = re.search(r"(1\d{8})", fn)
                            if m: ab = m.group(1)
                            with z.open(fi) as f:
                                wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
                                ws = wb.active
                                ic = str(ws.cell(20, 17).value or "").strip() == "입고금액"
                                cur = ""
                                for ri in range(1, ws.max_row + 1):
                                    bv = str(ws.cell(ri, 2).value or "").strip()
                                    if bv and bv != "None": cur = bv
                                    cv = str(ws.cell(ri, 3).value or "").strip()
                                    if cv.startswith("PL") or cv.startswith("880"):
                                        in_ = str(ws.cell(ri - 1, 3).value or "").strip()
                                        qv  = str(ws.cell(ri - 1, 8).value or "0").strip()
                                        ext.append({"a": ab, "b": cv, "loc": vmap.get(cur, "확인불가"), "item": in_, "qty": qv, "conf": ic})

            if not ext:
                st.warning("엑셀 파일에서 정보를 찾지 못했어요.")
            else:
                st.success(f"🎀 발주서 취합 완료! 총 {len(ext)}건")
                rows_html = "".join(
                    f'<tr class="{"wms-tr-even" if di%2==0 else "wms-tr-odd"}">'
                    f'<td class="wms-td" style="font-size:12px;">{rd["a"]}</td>'
                    f'<td class="wms-td" style="font-size:12px;">{rd["b"]}</td>'
                    f'<td class="wms-td">{rd["loc"]}</td>'
                    f'<td class="wms-td wms-td-left" style="{"color:var(--c-red);" if not rd["conf"] else ""}">{rd["item"]}</td>'
                    f'<td class="wms-td">{rd["qty"]}</td></tr>'
                    for di, rd in enumerate(ext)
                )
                st.markdown(f"""
                <div class="wms-card" style="padding:0;overflow:auto;">
                <table class="wms-table">
                  <thead><tr>
                    <th class="wms-th">발주서번호</th><th class="wms-th">상품바코드</th>
                    <th class="wms-th">로케이션</th><th class="wms-th">상품명</th><th class="wms-th">납품수량</th>
                  </tr></thead>
                  <tbody>{rows_html}</tbody>
                </table></div>
                """, unsafe_allow_html=True)

        except Exception as e:
            st.error(f"작업 중 문제가 발생했어요: {e}")


# ══════════════════════════════════════════════════════════════════
# 라우터
# ══════════════════════════════════════════════════════════════════
page = st.session_state.get("page", "login")

if page == "login" or not (svc_sa and svc_oauth):
    page_login()
elif page == "main":
    page_main()
elif page == "actual":
    page_actual()
elif page == "vendor":
    page_vendor()
elif page == "transfer":
    page_transfer()
elif page == "receive":
    page_receive()
elif page == "dispatch":
    page_dispatch()
elif page == "adjust":
    page_adjust()
elif page == "po":
    page_po()
